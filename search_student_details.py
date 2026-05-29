import openpyxl
import sys
import warnings
from datetime import datetime

# Suppress openpyxl formatting warnings
warnings.filterwarnings("ignore", category=UserWarning, module="openpyxl")

FILE_PATH = r"C:\Users\SriKriations\Favorites\Downloads\DAILY ACCOUNTS.xlsx"

def clean_str(val):
    if val is None:
        return ""
    return str(val).strip()

def normalize_class(c_str):
    c = clean_str(c_str).upper()
    c = "".join(c.split()) # remove spaces
    if not c:
        return ""
    
    # Standardize classes
    if c in ["NUR", "NURSERY", "N", "PLAY", "PLAYGROUP", "PG"]:
        return "NUR"
    if c in ["LKG", "PP-1", "PP1", "PPI", "L.KG", "PP-IA", "PPIA"]:
        return "LKG"
    if c in ["UKG", "PP-2", "PP2", "PPII", "U.KG", "PP-IIA", "PPIIA"]:
        return "UKG"
    
    # Standardize numerical classes
    for num, term in [("10", "10TH"), ("9", "9TH"), ("8", "8TH"), ("7", "7TH"), ("6", "6TH"), ("5", "5TH"), ("4", "4TH"), ("3", "3RD"), ("2", "2ND"), ("1", "1ST")]:
        if num in c:
            return term
            
    return c

COMMON_EXCLUDED_WORDS = {"REDDY", "GOUD", "KUMAR", "SINGH", "RAO", "DEVI", "KUMARI", "SHARMA", "SRI", "CHARY", "SREE", "CHARAN", "BAI", "LAL", "GOUD"}

def name_words_match(name1, name2):
    n1 = " ".join(clean_str(name1).upper().split())
    n2 = " ".join(clean_str(name2).upper().split())
    
    if not n1 or not n2:
        return False
        
    if n1 == n2:
        return True
        
    # Split into words
    words1 = n1.split()
    words2 = n2.split()
    
    # Split into single-letter initials and longer words
    initials1 = [w for w in words1 if len(w) == 1]
    initials2 = [w for w in words2 if len(w) == 1]
    
    long_words1 = [w for w in words1 if len(w) > 1 and w not in COMMON_EXCLUDED_WORDS]
    long_words2 = [w for w in words2 if len(w) > 1 and w not in COMMON_EXCLUDED_WORDS]
    
    # Core name match (must share at least one core word)
    core1 = set(long_words1)
    core2 = set(long_words2)
    
    shared_core = core1.intersection(core2)
    if not shared_core:
        return False
        
    # Validate initials: any initial in name1 must match the start of some word in name2
    for init in initials1:
        if not any(w.startswith(init) for w in words2):
            return False
            
    # Validate initials: any initial in name2 must match the start of some word in name1
    for init in initials2:
        if not any(w.startswith(init) for w in words1):
            return False
            
    return True

def should_merge_names(name1, name2):
    n1 = " ".join(clean_str(name1).upper().split())
    n2 = " ".join(clean_str(name2).upper().split())
    
    if not n1 or not n2:
        return False
        
    if n1 == n2 or n1 in n2 or n2 in n1:
        return True
        
    # Extract unique significant words
    words1 = set([w for w in n1.split() if len(w) > 2 and w not in COMMON_EXCLUDED_WORDS])
    words2 = set([w for w in n2.split() if len(w) > 2 and w not in COMMON_EXCLUDED_WORDS])
    
    # Conflicting surnames check
    diff1 = words1 - words2
    diff2 = words2 - words1
    
    if diff1 and diff2:
        return False
        
    # They must share at least one significant word
    return any(w in words2 for w in words1) or len(words1) == 0 or len(words2) == 0

def format_date(val):
    if val is None:
        return ""
    if isinstance(val, datetime):
        return val.strftime("%Y-%m-%d")
    return str(val).split()[0]

def clean_phone(val):
    if val is None:
        return ""
    s = str(val).strip()
    if s.endswith(".0"):
        s = s[:-2]
    return s

def clean_aadhar(val):
    if val is None:
        return ""
    s = str(val).strip()
    if s.endswith(".0"):
        s = s[:-2]
    return s

def load_master_directory(wb):
    if "Sheet14" not in wb.sheetnames:
        return []
    sheet = wb["Sheet14"]
    profiles = []
    for r in range(5, sheet.max_row + 1):
        name_val = clean_str(sheet.cell(row=r, column=4).value)
        if not name_val:
            continue
        profiles.append({
            "name": name_val,
            "class": clean_str(sheet.cell(row=r, column=5).value),
            "section": clean_str(sheet.cell(row=r, column=6).value),
            "branch": clean_str(sheet.cell(row=r, column=3).value).upper(),
            "id": clean_str(sheet.cell(row=r, column=2).value),
            "father": clean_str(sheet.cell(row=r, column=7).value),
            "mother": clean_str(sheet.cell(row=r, column=8).value),
            "f_phone": clean_phone(sheet.cell(row=r, column=9).value),
            "m_phone": clean_phone(sheet.cell(row=r, column=10).value),
            "aadhar": clean_aadhar(sheet.cell(row=r, column=15).value),
            "address": clean_str(sheet.cell(row=r, column=16).value),
            "source": f"Master Directory (Sheet14:R{r})",
            "bus_route": "",
            "pickup_point": "",
            "committed_fee": 0.0
        })
    return profiles

def load_bus_route_wise(wb):
    if "BUS ROUTE WISE" not in wb.sheetnames:
        return []
    sheet = wb["BUS ROUTE WISE"]
    profiles = []
    for r in range(8, sheet.max_row + 1):
        name_val = clean_str(sheet.cell(row=r, column=4).value)
        if not name_val:
            continue
        phone = clean_phone(sheet.cell(row=r, column=6).value)
        class_val = clean_str(sheet.cell(row=r, column=5).value)
        pickup = clean_str(sheet.cell(row=r, column=7).value)
        c_fee_val = sheet.cell(row=r, column=9).value
        
        c_fee = 0.0
        if c_fee_val is not None:
            try:
                c_fee = float(c_fee_val)
            except ValueError:
                pass
                
        profiles.append({
            "name": name_val,
            "class": class_val,
            "section": "",
            "branch": "MNB",
            "id": "",
            "father": "",
            "mother": "",
            "f_phone": phone,
            "m_phone": "",
            "aadhar": "",
            "address": "",
            "source": f"Bus Route Wise (R{r})",
            "bus_route": "MNB Route",
            "pickup_point": pickup,
            "committed_fee": c_fee
        })
    return profiles

def load_sheet17(wb):
    if "Sheet17" not in wb.sheetnames:
        return []
    sheet = wb["Sheet17"]
    profiles = []
    for r in range(6, sheet.max_row + 1):
        name_val = clean_str(sheet.cell(row=r, column=2).value)
        if not name_val:
            continue
        phone = clean_phone(sheet.cell(row=r, column=5).value)
        class_val = clean_str(sheet.cell(row=r, column=4).value)
        pickup = clean_str(sheet.cell(row=r, column=6).value)
        c_fee_val = sheet.cell(row=r, column=8).value
        
        c_fee = 0.0
        if c_fee_val is not None:
            try:
                c_fee = float(c_fee_val)
            except ValueError:
                pass
                
        profiles.append({
            "name": name_val,
            "class": class_val,
            "section": "",
            "branch": "MNB",
            "id": "",
            "father": "",
            "mother": "",
            "f_phone": phone,
            "m_phone": "",
            "aadhar": "",
            "address": "",
            "source": f"Sheet17 (R{r})",
            "bus_route": "MNB Route",
            "pickup_point": pickup,
            "committed_fee": c_fee
        })
    return profiles

def load_registers(wb, sheet_name, name_col, class_col, fee_col, start_row, default_branch):
    if sheet_name not in wb.sheetnames:
        return []
    sheet = wb[sheet_name]
    profiles = []
    for r in range(start_row, sheet.max_row + 1):
        name_val = clean_str(sheet.cell(row=r, column=name_col).value)
        name_upper = name_val.upper()
        if not name_val or any(x in name_upper for x in ["TOTAL", "GRAND TOTAL", "SUBTOTAL", "EXPENDITURE", "MONTHLY", "WEEK"]):
            break
            
        class_val = clean_str(sheet.cell(row=r, column=class_col).value)
        c_fee_val = sheet.cell(row=r, column=fee_col).value if fee_col else None
        
        c_fee = 0.0
        if c_fee_val is not None:
            try:
                c_fee = float(c_fee_val)
            except ValueError:
                pass
                
        profiles.append({
            "name": name_val,
            "class": class_val,
            "section": "",
            "branch": default_branch,
            "id": "",
            "father": "",
            "mother": "",
            "f_phone": "",
            "m_phone": "",
            "aadhar": "",
            "address": "",
            "source": f"{sheet_name} (R{r})",
            "bus_route": "",
            "pickup_point": "",
            "committed_fee": c_fee
        })
    return profiles

def merge_profiles(base, new_prof):
    if not base["id"] and new_prof["id"]: base["id"] = new_prof["id"]
    if not base["section"] and new_prof["section"]: base["section"] = new_prof["section"]
    if not base["father"] and new_prof["father"]: base["father"] = new_prof["father"]
    if not base["mother"] and new_prof["mother"]: base["mother"] = new_prof["mother"]
    if not base["f_phone"] and new_prof["f_phone"]: base["f_phone"] = new_prof["f_phone"]
    if not base["m_phone"] and new_prof["m_phone"]: base["m_phone"] = new_prof["m_phone"]
    if not base["aadhar"] and new_prof["aadhar"]: base["aadhar"] = new_prof["aadhar"]
    if not base["address"] and new_prof["address"]: base["address"] = new_prof["address"]
    if not base["bus_route"] and new_prof["bus_route"]: base["bus_route"] = new_prof["bus_route"]
    if not base["pickup_point"] and new_prof["pickup_point"]: base["pickup_point"] = new_prof["pickup_point"]
    if base["committed_fee"] == 0.0 and new_prof["committed_fee"] > 0.0:
        base["committed_fee"] = new_prof["committed_fee"]
    base["source"] += " + " + new_prof["source"]

def build_student_profiles(wb, all_payments):
    profiles = []
    
    # Load all roster sources
    profiles.extend(load_master_directory(wb))
    profiles.extend(load_bus_route_wise(wb))
    profiles.extend(load_sheet17(wb))
    profiles.extend(load_registers(wb, "Sheet19", 1, 2, 4, 5, "RCB"))
    profiles.extend(load_registers(wb, "Sheet18", 3, 4, 6, 17, "RCB"))
    profiles.extend(load_registers(wb, "Sheet20", 1, 2, 4, 7, "RCB"))
    
    # Extract unique student names + classes from all payments to act as fallback profiles
    for p in all_payments:
        sheet_clean = p["sheet"].upper()
        branch = "GLOBAL"
        if "RCB" in sheet_clean or sheet_clean == "RCB AC":
            branch = "RCB"
        elif "SNB" in sheet_clean or sheet_clean == "SNB AC":
            branch = "SNB"
        elif "MNB" in sheet_clean or sheet_clean == "MNB AC":
            branch = "MNB"
            
        profiles.append({
            "name": p["name"],
            "class": p["class"],
            "section": "",
            "branch": branch,
            "id": "",
            "father": "",
            "mother": "",
            "f_phone": "",
            "m_phone": "",
            "aadhar": "",
            "address": "",
            "source": f"Payment Sheet ({p['sheet']}:R{p['row']})",
            "bus_route": "",
            "pickup_point": "",
            "committed_fee": 0.0
        })
    
    # Deduplicate and merge profiles
    deduped = []
    for p in profiles:
        matched = False
        for d in deduped:
            # Check merge possibility
            # Match branch if not GLOBAL on either side, otherwise allow merge
            branch_match = (p["branch"] == d["branch"] or p["branch"] == "GLOBAL" or d["branch"] == "GLOBAL")
            
            if (should_merge_names(p["name"], d["name"]) and 
                normalize_class(p["class"]) == normalize_class(d["class"]) and
                branch_match):
                # Update branch to more specific if either is GLOBAL
                if d["branch"] == "GLOBAL" and p["branch"] != "GLOBAL":
                    d["branch"] = p["branch"]
                merge_profiles(d, p)
                matched = True
                break
        if not matched:
            deduped.append(p)
            
    return deduped

def get_payments(wb):
    payments = []
    
    for sheet_name in wb.sheetnames:
        if sheet_name in ["Sheet14", "Sheet19", "BUS ROUTE WISE", "Sheet18", "Sheet17", "Sheet20", "DIESEL", "BUS REPAIR", "MONTHLY", "BOOKS"]:
            continue
            
        sheet = wb[sheet_name]
        max_row = min(sheet.max_row, 100) # search first 100 rows for headers
        table_headers = []
        
        for r in range(1, max_row + 1):
            row_vals = [sheet.cell(row=r, column=c).value for c in range(1, sheet.max_column + 1)]
            
            for c_idx, val in enumerate(row_vals):
                c = c_idx + 1
                val_str = clean_str(val).upper()
                
                if ("NAME" in val_str or "STUDENT" in val_str) and "ABOUT" not in val_str and "EXPENDITURE" not in val_str:
                    name_col = c
                    date_col = None
                    class_col = None
                    amount_cols = []
                    
                    for other_c_idx, other_val in enumerate(row_vals):
                        other_c = other_c_idx + 1
                        other_val_str = clean_str(other_val).upper()
                        
                        if not other_val_str:
                            continue
                        
                        if "DATE" in other_val_str or other_val_str == "DT":
                            if date_col is None or abs(other_c - name_col) < abs(date_col - name_col):
                                date_col = other_c
                                
                        if "CLASS" in other_val_str or other_val_str == "CL":
                            if class_col is None or abs(other_c - name_col) < abs(class_col - name_col):
                                class_col = other_c
                                
                        if any(x in other_val_str for x in ["AMOUNT", "CASH", "ONLINE", "TOTAL", "FEE", "COLLECTED", "PAYMENT"]):
                            if other_c > name_col and (other_c - name_col) <= 6:
                                amount_cols.append((other_c, other_val_str))
                    
                    table_headers.append({
                        "header_row": r,
                        "name_col": name_col,
                        "date_col": date_col,
                        "class_col": class_col,
                        "amount_cols": amount_cols
                    })
                    
        for tb in table_headers:
            name_col = tb["name_col"]
            date_col = tb["date_col"]
            class_col = tb["class_col"]
            amount_cols = tb["amount_cols"]
            start_row = tb["header_row"] + 1
            
            for r in range(start_row, sheet.max_row + 1):
                name_val = clean_str(sheet.cell(row=r, column=name_col).value)
                name_upper = name_val.upper()
                
                if not name_val or any(x in name_upper for x in ["TOTAL", "GRAND TOTAL", "SUBTOTAL", "EXPENDITURE", "MONTHLY", "WEEK"]):
                    break
                    
                date_val = format_date(sheet.cell(row=r, column=date_col).value) if date_col else ""
                class_val = clean_str(sheet.cell(row=r, column=class_col).value) if class_col else ""
                
                pay_items = []
                for ac, label in amount_cols:
                    amt_val = sheet.cell(row=r, column=ac).value
                    if amt_val is not None:
                        try:
                            amt_num = float(amt_val)
                            if amt_num != 0:
                                pay_items.append((label, amt_num))
                        except ValueError:
                            if clean_str(amt_val):
                                pay_items.append((label, clean_str(amt_val)))
                                
                payments.append({
                    "sheet": sheet_name,
                    "date": date_val,
                    "name": name_val,
                    "class": class_val,
                    "payments": pay_items,
                    "row": r
                })
                
    return payments

def main():
    if len(sys.argv) < 2:
        print("Usage: python search_student_details.py <student_name>")
        sys.exit(1)
        
    query = " ".join(sys.argv[1:])
    print(f"\nScanning school database for: '{query}'...")
    
    wb = openpyxl.load_workbook(FILE_PATH, data_only=True)
    
    print("Retrieving all payment records...")
    all_payments = get_payments(wb)
    
    print("Building unified Student Registry from all sheets (including registers and payment fallbacks)...")
    unified_profiles = build_student_profiles(wb, all_payments)
    print(f"Loaded {len(unified_profiles)} unique student profiles.")
    
    # Find matching profiles
    matching_students = []
    for s in unified_profiles:
        if query.upper() in s["name"].upper() or name_words_match(s["name"], query):
            matching_students.append(s)
            
    if not matching_students:
        print(f"\nNo student profiles or payment records found matching '{query}'.")
        return
        
    print(f"Found {len(matching_students)} matching student(s) in Unified Registry.")
    
    # Correlate payments to each matched student profile
    for idx, student in enumerate(matching_students, 1):
        print(f"\n================================================================================")
        print(f"STUDENT #{idx}: {student['name'].upper()} ({student['branch']})")
        print(f"================================================================================")
        print(f"  ID           : {student['id'] if student['id'] else 'Not Found'}")
        print(f"  Class        : {student['class']} {('Section ' + student['section']) if student['section'] else ''}")
        print(f"  Aadhar No    : {student['aadhar'] if student['aadhar'] else 'Not Provided'}")
        print(f"  Father       : {student['father'] if student['father'] else 'Not Found'}")
        print(f"  Mother       : {student['mother'] if student['mother'] else 'Not Found'}")
        
        phones = []
        if student["f_phone"]: phones.append(f"Father/Primary: {student['f_phone']}")
        if student["m_phone"]: phones.append(f"Mother: {student['m_phone']}")
        print(f"  Phone Num(s) : {', '.join(phones) if phones else 'Not Provided'}")
        
        print(f"  Address      : {student['address'] if student['address'] else 'Not Provided'}")
        if student["bus_route"]:
            print(f"  Bus Route    : {student['bus_route']} | Pick-up: {student['pickup_point']}")
        if student["committed_fee"] > 0:
            print(f"  Committed Fee: Rs. {student['committed_fee']:,.2f}")
        print(f"  Source Sheet : {student['source']}")
        print(f"--------------------------------------------------------------------------------")
        
        student_payments = []
        for p in all_payments:
            name_match = name_words_match(p["name"], student["name"])
            
            sheet_clean = p["sheet"].upper()
            branch_match = False
            if student["branch"] == "GLOBAL" or student["branch"] in sheet_clean:
                branch_match = True
            elif student["branch"] == "RCB" and "RCB" in sheet_clean:
                branch_match = True
            elif student["branch"] == "SNB" and "SNB" in sheet_clean:
                branch_match = True
            elif student["branch"] == "MNB" and "MNB" in sheet_clean:
                branch_match = True
            elif sheet_clean in ["SHEET5"]: 
                branch_match = True
                
            class_match = True
            if p["class"] and student["class"]:
                norm_p_class = normalize_class(p["class"])
                norm_s_class = normalize_class(student["class"])
                if norm_p_class and norm_s_class and norm_p_class != norm_s_class:
                    class_match = False
                    
            if name_match and branch_match and class_match:
                student_payments.append(p)
                
        if not student_payments:
            print("  No payment records found for this student.")
        else:
            print(f"  Found {len(student_payments)} payment record(s):")
            print(f"    {'Sheet':<10} | {'Date':<10} | {'Name in Sheet':<25} | {'Class':<6} | {'Payments Recorded'}")
            print(f"    " + "-" * 80)
            subtotal = 0.0
            for sp in student_payments:
                pays = ", ".join([f"Rs. {val:,.2f} ({label})" if isinstance(val, (int, float)) else f"{val} ({label})" for label, val in sp["payments"]])
                for label, val in sp["payments"]:
                    if isinstance(val, (int, float)):
                        subtotal += val
                print(f"    {sp['sheet']:<10} | {sp['date']:<10} | {sp['name']:<25} | {sp['class']:<6} | {pays}")
            print(f"    " + "-" * 80)
            print(f"    Total Payments Extracted: Rs. {subtotal:,.2f}")
            
    # Print unmatched payments matching query name to make sure nothing was missed
    print(f"\n================================================================================")
    print("UNMATCHED / RAW SEARCH CHECK (Double-Checking all instances of query in payments)")
    print(f"================================================================================")
    raw_matches = [p for p in all_payments if query.upper() in p["name"].upper()]
    if not raw_matches:
        print("No matches in raw sheets outside the master profiles.")
    else:
        print(f"Found {len(raw_matches)} occurrences of '{query}' in payment sheets:")
        for rm in raw_matches:
            pays = ", ".join([f"Rs. {val:,.2f} ({label})" if isinstance(val, (int, float)) else f"{val} ({label})" for label, val in rm["payments"]])
            print(f"  Sheet: {rm['sheet']:<8} | Date: {rm['date']:<10} | Name: {rm['name']:<25} | Class: {rm['class']:<6} | {pays}")

if __name__ == "__main__":
    main()
