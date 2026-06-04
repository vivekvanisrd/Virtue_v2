import json
import openpyxl
import re

# Load student_data.json
JSON_PATH = "public/student-directory/student_data.json"
with open(JSON_PATH, "r", encoding="utf-8") as f:
    student_profiles = json.load(f)

# Build a mapping of normalized student names to their branch
student_branches = {}
for s in student_profiles:
    norm_name = re.sub(r"[\s\.]", "", s["name"]).upper()
    student_branches[norm_name] = s["branch"]

# Load Excel workbook
EXCEL_PATH = r"E:\accounts\sorted fee 2025 2026.xlsx"
wb = openpyxl.load_workbook(EXCEL_PATH, read_only=True)
sheet = wb["Ucban"]

branch_map_counts = {}
for r in range(4, sheet.max_row + 1):
    name = sheet.cell(row=r, column=8).value
    if not name:
        continue
    name_str = str(name).strip()
    norm_name = re.sub(r"[\s\.]", "", name_str).upper()
    
    if norm_name in student_branches:
        mapped_branch = student_branches[norm_name]
    else:
        mapped_branch = "UNMAPPED"
        
    branch_map_counts[mapped_branch] = branch_map_counts.get(mapped_branch, 0) + 1

print("Mapping of Excel students to dashboard branches:")
for branch, count in sorted(branch_map_counts.items(), key=lambda x: x[1], reverse=True):
    print(f"Branch {branch}: {count} rows")
