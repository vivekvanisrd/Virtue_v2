import openpyxl
import sys

sys.stdout.reconfigure(encoding='utf-8')

wb = openpyxl.load_workbook(r"E:\accounts\sorted fee 2025 2026.xlsx", data_only=True)
sheet = wb['Feb_Mar_Apr_2026']

print("=== DECEMBER 2025 ROWS (773-869) ===")
for r in range(773, 870):
    vals = [sheet.cell(row=r, column=c).value for c in range(1, 13)]
    formatted_vals = [val.strftime('%Y-%m-%d') if hasattr(val, 'strftime') else val for val in vals]
    print(f"Row {r:3d}: {formatted_vals}")

print("\n=== JANUARY 2026 ROWS (870-930) ===")
for r in range(870, 931):
    vals = [sheet.cell(row=r, column=c).value for c in range(1, 13)]
    formatted_vals = [val.strftime('%Y-%m-%d') if hasattr(val, 'strftime') else val for val in vals]
    print(f"Row {r:3d}: {formatted_vals}")

wb.close()
