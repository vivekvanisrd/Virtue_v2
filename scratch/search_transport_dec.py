import openpyxl
import sys

sys.stdout.reconfigure(encoding='utf-8')

wb = openpyxl.load_workbook(r"E:\accounts\sorted fee 2025 2026.xlsx", data_only=True)
sheet = wb['Feb_Mar_Apr_2026']

print("Searching for receipts 5197-5205:")
for r in range(3, sheet.max_row + 1):
    rec_val = sheet.cell(row=r, column=6).value
    if rec_val is not None:
        try:
            rec_int = int(float(str(rec_val).strip()))
            if 5197 <= rec_int <= 5205:
                vals = [sheet.cell(row=r, column=c).value for c in range(1, 13)]
                print(f"Row {r:4d}: {vals}")
        except ValueError:
            pass

wb.close()
