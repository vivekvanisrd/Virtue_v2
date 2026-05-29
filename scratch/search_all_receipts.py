import openpyxl
import sys

sys.stdout.reconfigure(encoding='utf-8')

wb = openpyxl.load_workbook(r"E:\accounts\sorted fee 2025 2026.xlsx", data_only=True)

# Search both sheets
for sheet_name in wb.sheetnames:
    sheet = wb[sheet_name]
    print(f"\nSearching in sheet: {sheet_name}")
    found_count = 0
    for r in range(1, sheet.max_row + 1):
        for c in range(1, sheet.max_column + 1):
            val = sheet.cell(row=r, column=c).value
            if val is not None:
                val_str = str(val).strip()
                # Check if it contains any of our receipts
                for rec in range(5282, 5301):
                    if val_str == str(rec) or val_str == f"{rec}.0":
                        print(f"  Row {r}, Col {c} (ColName {sheet.cell(row=2, column=c).value if sheet.max_row >= 2 else ''}): {val_str} | Full Row: {[sheet.cell(row=r, column=col).value for col in range(1, min(15, sheet.max_column+1))]}")
                        found_count += 1
    print(f"Found {found_count} occurrences in {sheet_name}")

wb.close()
