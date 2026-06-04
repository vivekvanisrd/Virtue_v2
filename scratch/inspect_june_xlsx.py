import openpyxl
import sys

sys.stdout.reconfigure(encoding='utf-8')
wb = openpyxl.load_workbook(r"E:\accounts\RCB june 2025 26.xlsx", data_only=True)
print("Sheets in RCB june 2025 26.xlsx:")
print(wb.sheetnames)

for name in wb.sheetnames:
    sheet = wb[name]
    print(f"\nSheet {name}: {sheet.max_row} rows, {sheet.max_column} cols")
    for r in range(1, min(6, sheet.max_row + 1)):
        print(f"  Row {r}: {[sheet.cell(row=r, column=c).value for c in range(1, min(10, sheet.max_column+1))]}")

wb.close()
