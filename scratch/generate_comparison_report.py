import json
import re
import openpyxl
from datetime import datetime

# Load student_data.json
JSON_PATH = "public/student-directory/student_data.json"
with open(JSON_PATH, "r", encoding="utf-8") as f:
    student_profiles = json.load(f)

# Extract only PDF-based payments from student_data.json
json_payments = {}
for profile in student_profiles:
    for tr in profile.get("payments", []):
        sheet = tr.get("sheet", "")
        if not sheet.lower().endswith(".pdf"):
            continue
            
        row_num = tr.get("row", "")
        date = tr.get("date", "")
        receipt_no = tr.get("receipt_no", "")
        name_in_sheet = tr.get("name", "")
        
        cash_amt = 0.0
        online_amt = 0.0
        pay_details = []
        for label, val in tr.get("payments", []):
            try:
                amt = float(val)
            except (ValueError, TypeError):
                continue
            pay_details.append((label, amt))
            lbl_upper = str(label).upper()
            if "CASH" in lbl_upper:
                cash_amt += amt
            else:
                online_amt += amt
                
        tx_key = (sheet, date, name_in_sheet, receipt_no, str(sorted(pay_details)), row_num)
        json_payments[tx_key] = {
            "student_name": profile["name"],
            "name_in_sheet": name_in_sheet,
            "class": profile["class"],
            "class_in_sheet": tr.get("class", ""),
            "branch": profile["branch"],
            "date": date,
            "receipt_no": receipt_no,
            "cash": cash_amt,
            "online": online_amt,
            "total": cash_amt + online_amt,
            "sheet": sheet,
            "row": row_num
        }

json_list = list(json_payments.values())

# Load Excel workbook
EXCEL_PATH = r"E:\accounts\sorted fee 2025 2026.xlsx"
wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True)
sheet = wb["Ucban"]

def normalize_name(name):
    if not name:
        return ""
    return re.sub(r"[\s\.]", "", str(name)).upper()

def normalize_receipt(rec):
    if rec is None:
        return ""
    rec_str = str(rec).strip()
    if rec_str.endswith(".0"):
        rec_str = rec_str[:-2]
    rec_str = re.sub(r"\s+", "", rec_str)
    if rec_str.isdigit():
        return str(int(rec_str))
    return rec_str.upper()

def parse_date(val):
    if val is None:
        return ""
    if isinstance(val, datetime):
        return val.strftime("%Y-%m-%d")
    val_str = str(val).strip()
    for fmt in ("%d/%m/%Y", "%Y-%m-%d", "%d-%m-%Y"):
        try:
            return datetime.strptime(val_str, fmt).strftime("%Y-%m-%d")
        except ValueError:
            pass
    return val_str.split()[0]

excel_payments = []
for r in range(4, sheet.max_row + 1):
    name = sheet.cell(row=r, column=8).value
    if not name:
        continue
    name_str = str(name).upper()
    if any(x in name_str for x in ["DEPOSIT", "EXPENDITURE", "TOTAL", "GRAND TOTAL", "CONTRA", "CEMENT", "DIESEL"]):
        continue
        
    date_val = sheet.cell(row=r, column=6).value
    receipt_val = sheet.cell(row=r, column=7).value
    cash_val = sheet.cell(row=r, column=10).value
    online_val = sheet.cell(row=r, column=11).value
    total_val = sheet.cell(row=r, column=12).value
    
    cash = float(cash_val) if cash_val is not None else 0.0
    online = float(online_val) if online_val is not None else 0.0
    total = float(total_val) if total_val is not None else (cash + online)
    
    excel_payments.append({
        "row": r,
        "student_name": str(name).strip(),
        "class": str(sheet.cell(row=r, column=9).value).strip() if sheet.cell(row=r, column=9).value else "",
        "date": parse_date(date_val),
        "receipt_no": normalize_receipt(receipt_val),
        "cash": cash,
        "online": online,
        "total": total,
        "branch": sheet.cell(row=r, column=4).value or ""
    })

# Match logic
matched_json_indices = set()
unmatched_excel = []
unmatched_json = []

for ep in excel_payments:
    norm_name = normalize_name(ep["student_name"])
    norm_receipt = normalize_receipt(ep["receipt_no"])
    ep_date = ep["date"]
    
    found_match = False
    
    # 1. Match Name + Receipt + Date + Total
    for idx, jp in enumerate(json_list):
        if idx in matched_json_indices:
            continue
        name_match = normalize_name(jp["name_in_sheet"]) == norm_name or normalize_name(jp["student_name"]) == norm_name
        receipt_match = normalize_receipt(jp["receipt_no"]) == norm_receipt if norm_receipt else True
        date_match = jp["date"] == ep_date if ep_date and jp["date"] else True
        amount_match = abs(jp["total"] - ep["total"]) < 0.01
        
        if name_match and receipt_match and date_match and amount_match:
            matched_json_indices.add(idx)
            found_match = True
            break
            
    if not found_match:
        # 2. Match Name + Amount + (Date or Receipt)
        for idx, jp in enumerate(json_list):
            if idx in matched_json_indices:
                continue
            name_match = normalize_name(jp["name_in_sheet"]) == norm_name or normalize_name(jp["student_name"]) == norm_name
            receipt_match = normalize_receipt(jp["receipt_no"]) == norm_receipt
            date_match = jp["date"] == ep_date
            amount_match = abs(jp["total"] - ep["total"]) < 0.01
            
            if name_match and amount_match and (date_match or receipt_match):
                matched_json_indices.add(idx)
                found_match = True
                break
                
    if not found_match:
        unmatched_excel.append(ep)

for idx, jp in enumerate(json_list):
    if idx not in matched_json_indices:
        unmatched_json.append(jp)

# Categorize unmatched records
detail_mismatches = []
excel_unmatched_remaining = []

for ep in unmatched_excel:
    ep_name_norm = normalize_name(ep["student_name"])
    matching_jps = [jp for jp in unmatched_json if normalize_name(jp["name_in_sheet"]) == ep_name_norm or normalize_name(jp["student_name"]) == ep_name_norm]
    
    if matching_jps:
        detail_mismatches.append((ep, matching_jps))
    else:
        excel_unmatched_remaining.append(ep)

def share_words(name1, name2):
    n1 = re.sub(r"[^\w\s]", " ", name1.upper())
    n2 = re.sub(r"[^\w\s]", " ", name2.upper())
    words1 = set(w for w in n1.split() if len(w) > 2)
    words2 = set(w for w in n2.split() if len(w) > 2)
    return len(words1.intersection(words2)) > 0

spelling_variations = []
excel_unmatched_final = []

for ep in excel_unmatched_remaining:
    ep_name = ep["student_name"]
    matching_jps = [jp for jp in unmatched_json if share_words(jp["name_in_sheet"], ep_name) or share_words(jp["student_name"], ep_name)]
    if matching_jps:
        spelling_variations.append((ep, matching_jps))
    else:
        excel_unmatched_final.append(ep)

json_unmatched_remaining = []
for jp in unmatched_json:
    jp_name = jp["name_in_sheet"]
    if not any(normalize_name(ep["student_name"]) == normalize_name(jp_name) for ep in unmatched_excel):
        if not any(share_words(ep["student_name"], jp_name) for ep in unmatched_excel):
            json_unmatched_remaining.append(jp)

# Create markdown report
REPORT_PATH = r"C:\Users\SriKriations\.gemini\antigravity\brain\9c7156a0-e1c4-462b-a3f8-70303075f29e\fee_audit_comparison_report.md"

with open(REPORT_PATH, "w", encoding="utf-8") as f:
    f.write("# Fee Registry Audit: Differences Report\n\n")
    f.write("This report presents the differences between the local student directory dashboard database (`student_data.json`) and the sorted registry Excel spreadsheet `E:\\accounts\\sorted fee 2025 2026.xlsx` (sheet `Ucban`).\n\n")
    
    # Summary Table
    f.write("## Executive Summary\n\n")
    f.write("| Dataset | Total Records | Matched | Unmatched |\n")
    f.write("| --- | --- | --- | --- |\n")
    f.write(f"| **Dashboard (PDF Sourced)** | {len(json_list)} | {len(matched_json_indices)} | {len(unmatched_json)} |\n")
    f.write(f"| **Excel Sheet (Ucban)** | {len(excel_payments)} | {len(matched_json_indices)} | {len(unmatched_excel)} |\n\n")
    
    f.write("> [!NOTE]\n")
    f.write("> The dashboard includes all payments from the 2025-2026 academic year compiled from the register PDFs, whereas `sorted fee 2025 2026.xlsx` is a manually/semi-manually digitized spreadsheet from the registers. Mismatches highlight transcription variations, spelling differences, split transaction issues, and missing entries.\n\n")
    
    # 1. Detail Mismatches
    f.write("## 1. Student Exists in Both, but Details Mismatch\n")
    f.write("These records have the exact same student name in both the Excel sheet and the dashboard, but their transaction details (date, receipt number, or paid amount) differ. E.g., split payments in the Excel sheet vs combined or different amounts in the dashboard.\n\n")
    
    f.write("| Student Name | Excel Row | Excel Date | Excel Receipt | Excel Total | JSON Date | JSON Receipt | JSON Total | Source Sheet |\n")
    f.write("| --- | --- | --- | --- | --- | --- | --- | --- | --- |\n")
    for ep, jps in detail_mismatches:
        for jp in jps:
            f.write(f"| {ep['student_name']} | Row {ep['row']} | {ep['date']} | {ep['receipt_no']} | Rs. {ep['total']:,} | {jp['date']} | {jp['receipt_no']} | Rs. {jp['total']:,} | {jp['sheet']} |\n")
    f.write("\n")
    
    # 2. Spelling Variations
    f.write("## 2. Spelling Variations & Partial Matches\n")
    f.write("These records show student names that match partially or have spelling variations (such as initials flipped or different phonetic transliteration), resulting in matching student identities but failing exact name matching.\n\n")
    
    f.write("| Excel Student Name (Row) | Excel Details | JSON Student Name | JSON Details | Source Sheet |\n")
    f.write("| --- | --- | --- | --- | --- |\n")
    for ep, jps in spelling_variations:
        excel_details = f"Date: {ep['date']} \| Rec: {ep['receipt_no']} \| Rs. {ep['total']:,}"
        for jp in jps[:1]:  # Show top match
            json_details = f"Date: {jp['date']} \| Rec: {jp['receipt_no']} \| Rs. {jp['total']:,}"
            f.write(f"| {ep['student_name']} (Row {ep['row']}) | {excel_details} | {jp['name_in_sheet']} | {json_details} | {jp['sheet']} |\n")
    f.write("\n")
    
    # 3. Only in Excel
    f.write("## 3. Present in Excel, Completely Missing in Dashboard\n")
    f.write("These transactions are listed in the Excel spreadsheet but are completely absent from the dashboard (no matches by name, receipt, or other indicators in the PDF records).\n\n")
    
    f.write("| Excel Row | Student Name | Class | Date | Receipt No | Cash | Online | Total |\n")
    f.write("| --- | --- | --- | --- | --- | --- | --- | --- |\n")
    for ep in excel_unmatched_final:
        f.write(f"| Row {ep['row']} | {ep['student_name']} | {ep['class']} | {ep['date']} | {ep['receipt_no']} | Rs. {ep['cash']:,} | Rs. {ep['online']:,} | Rs. {ep['total']:,} |\n")
    f.write("\n")
    
    # 4. Only in Dashboard
    f.write("## 4. Present in Dashboard, Completely Missing in Excel\n")
    f.write("These transactions were transcribed from the PDF register copies and exist in the student database, but are completely absent from the Excel spreadsheet.\n\n")
    
    f.write("| Source Sheet | Row | Student Name | Class | Date | Receipt No | Total Paid |\n")
    f.write("| --- | --- | --- | --- | --- | --- | --- |\n")
    for jp in json_unmatched_remaining:
        f.write(f"| {jp['sheet']} | Row {jp['row']} | {jp['name_in_sheet']} | {jp['class_in_sheet']} | {jp['date']} | {jp['receipt_no']} | Rs. {jp['total']:,} |\n")
    f.write("\n")

print("Report written successfully to", REPORT_PATH)
