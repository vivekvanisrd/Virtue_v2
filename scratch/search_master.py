import openpyxl
import sys
import warnings

# Suppress openpyxl warnings
warnings.filterwarnings("ignore", category=UserWarning, module="openpyxl")

FILE_PATH = r"C:\Users\SriKriations\Favorites\Downloads\DAILY ACCOUNTS.xlsx"

def clean_str(val):
    if val is None:
        return ""
    return str(val).strip()

def search_master(query):
    wb = openpyxl.load_workbook(FILE_PATH, data_only=True)
    if "Sheet14" not in wb.sheetnames:
        print("Sheet14 not found in workbook.")
        return
        
    sheet = wb["Sheet14"]
    results = []
    
    # We will look for headers in row 1-10 to find student name, father name, etc.
    header_row = 5 # from our inspection output
    headers = [clean_str(sheet.cell(row=header_row, column=c).value).upper() for c in range(1, sheet.max_column + 1)]
    print("Master Headers:", headers)
    
    for r in range(header_row + 1, sheet.max_row + 1):
        name_val = clean_str(sheet.cell(row=r, column=4).value) # column 4 is Student Name
        if not name_val:
            continue
            
        if query.upper() in name_val.upper():
            # Extract details
            row_data = {
                "id": clean_str(sheet.cell(row=r, column=2).value),
                "branch": clean_str(sheet.cell(row=r, column=3).value),
                "name": name_val,
                "class": clean_str(sheet.cell(row=r, column=5).value),
                "section": clean_str(sheet.cell(row=r, column=6).value),
                "father": clean_str(sheet.cell(row=r, column=7).value),
                "mother": clean_str(sheet.cell(row=r, column=8).value),
                "f_phone": clean_str(sheet.cell(row=r, column=9).value),
                "m_phone": clean_str(sheet.cell(row=r, column=10).value),
                "aadhar": clean_str(sheet.cell(row=r, column=15).value),
                "address": clean_str(sheet.cell(row=r, column=16).value)
            }
            results.append(row_data)
            
    return results

if __name__ == "__main__":
    q = " ".join(sys.argv[1:]) if len(sys.argv) > 1 else "ANWIKA"
    print(f"Searching master directory for: '{q}'...")
    res = search_master(q)
    print(f"Found {len(res)} matches:")
    for r in res:
        print(f"\nName: {r['name']} ({r['branch']})")
        print(f"  ID: {r['id']} | Class: {r['class']} {r['section']}")
        print(f"  Parents: Father: {r['father']} (Phone: {r['f_phone']}) | Mother: {r['mother']} (Phone: {r['m_phone']})")
        print(f"  Aadhar: {r['aadhar']}")
        print(f"  Address: {r['address']}")
