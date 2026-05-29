import openpyxl
import re

file_path = r"C:\Users\SriKriations\Favorites\Downloads\DAILY ACCOUNTS.xlsx"
wb = openpyxl.load_workbook(file_path, data_only=True)

def clean_str(val):
    if val is None:
        return ""
    return str(val).strip().upper()

for sheet_name in wb.sheetnames:
    sheet = wb[sheet_name]
    print(f"\n======================================")
    print(f"SHEET: {sheet_name}")
    print(f"======================================")
    
    # Scan first 50 rows for headers
    max_row = min(sheet.max_row, 50)
    for r in range(1, max_row + 1):
        row_vals = [sheet.cell(row=r, column=c).value for c in range(1, sheet.max_column + 1)]
        
        # Look for cells containing "NAME" or "STUDENT"
        for c_idx, val in enumerate(row_vals):
            c = c_idx + 1 # 1-based column index
            val_str = clean_str(val)
            
            # Match "NAME" or "STUDENT" but avoid headers that represent other things
            if ("NAME" in val_str or "STUDENT" in val_str) and "ABOUT" not in val_str and "EXPENDITURE" not in val_str:
                # We found a potential Name column! Let's locate Date, Class, and Amount columns in the same row.
                name_col = c
                date_col = None
                class_col = None
                amount_cols = []
                
                # Look for Date, Class, Amount in the vicinity (columns 1 to sheet.max_column)
                for other_c_idx, other_val in enumerate(row_vals):
                    other_c = other_c_idx + 1
                    other_val_str = clean_str(other_val)
                    
                    if not other_val_str:
                        continue
                    
                    # Date check
                    if "DATE" in other_val_str or "DT" == other_val_str:
                        # Find the Date column closest to the Name column
                        if date_col is None or abs(other_c - name_col) < abs(date_col - name_col):
                            date_col = other_c
                    
                    # Class check
                    if "CLASS" in other_val_str or other_val_str == "CL":
                        if class_col is None or abs(other_c - name_col) < abs(class_col - name_col):
                            class_col = other_c
                            
                    # Amount check (can have multiple amount columns, e.g. CASH, ONLINE, AMOUNT)
                    if any(x in other_val_str for x in ["AMOUNT", "CASH", "ONLINE", "TOTAL", "FEE", "COLLECTED", "PAYMENT"]):
                        # Only include amount columns that are close to the name column (e.g. within 6 columns)
                        if abs(other_c - name_col) <= 6 and other_c != name_col:
                            amount_cols.append((other_c, other_val_str))
                
                print(f"Row {r}, Col {name_col} ('{val_str}'):")
                print(f"  Detected Date Col: {date_col} ({clean_str(sheet.cell(row=r, column=date_col).value) if date_col else 'None'})")
                print(f"  Detected Class Col: {class_col} ({clean_str(sheet.cell(row=r, column=class_col).value) if class_col else 'None'})")
                print(f"  Detected Amount Cols: {[(ac, label) for ac, label in amount_cols]}")
