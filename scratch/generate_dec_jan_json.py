import openpyxl
import json
import os
import sys
import datetime

sys.stdout.reconfigure(encoding='utf-8')

# Source Excel
excel_path = r"E:\accounts\sorted fee 2025 2026.xlsx"
output_dir = r"J:\virtue_fb\virtue-v2\scratch\transcribed_json"
os.makedirs(output_dir, exist_ok=True)

wb = openpyxl.load_workbook(excel_path, data_only=True)
sheet = wb['Feb_Mar_Apr_2026']

def clean_name(name):
    if name is None:
        return ""
    name_str = str(name).strip().upper()
    # Replace double spaces with single
    while "  " in name_str:
        name_str = name_str.replace("  ", " ")
    return name_str

def clean_class(cls):
    if cls is None:
        return ""
    cls_str = str(cls).strip().upper()
    # Convert standard formats if needed (e.g., LKG, UKG, NUR)
    # 2ND -> 2ND, 6TH -> 6TH, etc.
    return cls_str

def parse_date(date_val):
    if isinstance(date_val, datetime.datetime):
        return date_val.strftime('%Y-%m-%d')
    if date_val is None:
        return ""
    # Try parsing string like DD/MM/YYYY
    d_str = str(date_val).strip()
    for fmt in ('%d/%m/%Y', '%Y-%m-%d', '%d-%m-%Y', '%d/%m/%y'):
        try:
            return datetime.datetime.strptime(d_str, fmt).strftime('%Y-%m-%d')
        except ValueError:
            pass
    return d_str

def make_payments(cash_val, online_val, remarks):
    payments = []
    cash = 0
    online = 0
    try:
        if cash_val:
            cash = int(float(str(cash_val).strip()))
    except ValueError:
        pass
    try:
        if online_val:
            online = int(float(str(online_val).strip()))
    except ValueError:
        pass
        
    remarks_lower = str(remarks).lower() if remarks else ""
    is_tf = "tf" in remarks_lower or "t/f" in remarks_lower or "term" in remarks_lower
    is_transport = "transport" in remarks_lower
    
    if cash > 0:
        if is_transport:
            payments.append(["CASH (TRANSPORT)", cash])
        elif is_tf:
            payments.append(["CASH (T.F)", cash])
        else:
            payments.append(["CASH", cash])
            
    if online > 0:
        if is_transport:
            payments.append(["ONLINE (TRANSPORT)", online])
        elif is_tf:
            payments.append(["ONLINE (T.F)", online])
        else:
            payments.append(["ONLINE", online])
            
    return payments

# Load row values helper
def get_row_data(row_idx):
    date_str = parse_date(sheet.cell(row=row_idx, column=5).value)
    name = clean_name(sheet.cell(row=row_idx, column=7).value)
    cls = clean_class(sheet.cell(row=row_idx, column=8).value)
    rec = str(sheet.cell(row=row_idx, column=6).value or "").strip()
    # Remove float decimal if present
    if rec.endswith(".0"):
        rec = rec[:-2]
    
    cash = sheet.cell(row=row_idx, column=9).value
    online = sheet.cell(row=row_idx, column=10).value
    remarks = sheet.cell(row=row_idx, column=12).value
    
    payments = make_payments(cash, online, remarks)
    
    return {
        "date": date_str,
        "name": name,
        "class": cls,
        "receipt_no": rec,
        "payments": payments,
        "sheet": "RCB Dec Jan 2025 26.pdf"
    }

# --- PAGE 0 ---
# Rows 773-790 from Excel
page_0 = []
for r in range(773, 791):
    page_0.append(get_row_data(r))

# Plus manually added transport rows
p0_transport = [
    {"date": "2025-11-29", "name": "V. OMKAR REDDY", "class": "VIII", "receipt_no": "5197", "payments": [["ONLINE (TRANSPORT)", 5000]], "sheet": "RCB Dec Jan 2025 26.pdf"},
    {"date": "2025-11-29", "name": "V. BHARGAV REDDY", "class": "VI", "receipt_no": "5198", "payments": [["ONLINE (TRANSPORT)", 5000]], "sheet": "RCB Dec Jan 2025 26.pdf"},
    {"date": "2025-12-03", "name": "K. RUGVIJA SREE", "class": "III", "receipt_no": "5199", "payments": [["ONLINE (TRANSPORT)", 4000]], "sheet": "RCB Dec Jan 2025 26.pdf"},
    {"date": "2025-12-04", "name": "G. DHANA SRI", "class": "V", "receipt_no": "5200", "payments": [["CASH (TRANSPORT)", 1000]], "sheet": "RCB Dec Jan 2025 26.pdf"},
    {"date": "2025-12-04", "name": "G. HANEESH", "class": "IV", "receipt_no": "5201", "payments": [["CASH (TRANSPORT)", 1000]], "sheet": "RCB Dec Jan 2025 26.pdf"},
    {"date": "2025-12-04", "name": "C. DIYASHARVANI", "class": "III", "receipt_no": "5202", "payments": [["ONLINE (TRANSPORT)", 5000]], "sheet": "RCB Dec Jan 2025 26.pdf"},
    {"date": "2025-12-04", "name": "ANUPAM JENA", "class": "3RD", "receipt_no": "5203", "payments": [["ONLINE (TRANSPORT)", 3000]], "sheet": "RCB Dec Jan 2025 26.pdf"},
    {"date": "2025-12-05", "name": "B. AARUSH", "class": "III", "receipt_no": "5204", "payments": [["CASH (TRANSPORT)", 2000]], "sheet": "RCB Dec Jan 2025 26.pdf"},
    {"date": "2025-12-05", "name": "B. AAKARSH", "class": "I", "receipt_no": "5205", "payments": [["CASH (TRANSPORT)", 2000]], "sheet": "RCB Dec Jan 2025 26.pdf"}
]
page_0.extend(p0_transport)

for idx, tx in enumerate(page_0):
    tx["row"] = idx + 1

# --- PAGE 1 ---
# Rows 791-814 from Excel
page_1 = []
for r in range(791, 815):
    page_1.append(get_row_data(r))

for idx, tx in enumerate(page_1):
    tx["row"] = idx + 1

# --- PAGE 2 ---
# Rows 815-840 from Excel
page_2 = []
for r in range(815, 841):
    page_2.append(get_row_data(r))

# Plus transport rows
p2_transport = [
    {"date": "2025-12-13", "name": "J. GOWTHAM", "class": "7TH", "receipt_no": "5206", "payments": [["CASH (TRANSPORT)", 1000]], "sheet": "RCB Dec Jan 2025 26.pdf"},
    {"date": "2025-12-13", "name": "JOSHIKA", "class": "V", "receipt_no": "5207", "payments": [["CASH (TRANSPORT)", 5000]], "sheet": "RCB Dec Jan 2025 26.pdf"}
]
page_2.extend(p2_transport)

for idx, tx in enumerate(page_2):
    tx["row"] = idx + 1

# --- PAGE 3 ---
# Rows 841-869 from Excel
page_3 = []
for r in range(841, 870):
    page_3.append(get_row_data(r))

for idx, tx in enumerate(page_3):
    tx["row"] = idx + 1

# --- PAGE 4 ---
page_4 = []

# --- PAGE 5 ---
# Rows 870-896 from Excel
page_5 = []
for r in range(870, 897):
    page_5.append(get_row_data(r))

for idx, tx in enumerate(page_5):
    tx["row"] = idx + 1

# --- PAGE 6 ---
# Manually added transport rows, then student rows (897-908)
page_6 = []
p6_transport = [
    {"date": "2025-12-30", "name": "J. GOWTHAM", "class": "VI", "receipt_no": "5206", "payments": [["CASH (TRANSPORT)", 1000]], "sheet": "RCB Dec Jan 2025 26.pdf"},
    {"date": "2025-12-30", "name": "JOSHIKA", "class": "V", "receipt_no": "5207", "payments": [["CASH (TRANSPORT)", 5000]], "sheet": "RCB Dec Jan 2025 26.pdf"},
    {"date": "2025-12-30", "name": "T. MANOJ", "class": "V", "receipt_no": "5208", "payments": [["CASH (TRANSPORT)", 4000]], "sheet": "RCB Dec Jan 2025 26.pdf"},
    {"date": "2025-12-30", "name": "T. HARINI", "class": "II", "receipt_no": "5209", "payments": [["CASH (TRANSPORT)", 4000]], "sheet": "RCB Dec Jan 2025 26.pdf"},
    {"date": "2026-01-02", "name": "G. LASYA PRIYA", "class": "IV", "receipt_no": "5210", "payments": [["ONLINE (TRANSPORT)", 3000]], "sheet": "RCB Dec Jan 2025 26.pdf"},
    {"date": "2026-01-02", "name": "G. NANDA KISHORE", "class": "II", "receipt_no": "5211", "payments": [["ONLINE (TRANSPORT)", 3000]], "sheet": "RCB Dec Jan 2025 26.pdf"},
    {"date": "2026-01-02", "name": "C. SAANVI", "class": "IV", "receipt_no": "5212", "payments": [["ONLINE (TRANSPORT)", 4000]], "sheet": "RCB Dec Jan 2025 26.pdf"},
    {"date": "2026-01-02", "name": "K. ABHIJEET GOUD", "class": "III", "receipt_no": "5213", "payments": [["ONLINE (TRANSPORT)", 4000]], "sheet": "RCB Dec Jan 2025 26.pdf"},
    {"date": "2026-01-03", "name": "G. SAI KARTHIK", "class": "II", "receipt_no": "5214", "payments": [["ONLINE (TRANSPORT)", 5000]], "sheet": "RCB Dec Jan 2025 26.pdf"},
    {"date": "2026-01-03", "name": "G. SAI THANVIK", "class": "III", "receipt_no": "5215", "payments": [["ONLINE (TRANSPORT)", 5000]], "sheet": "RCB Dec Jan 2025 26.pdf"},
    {"date": "2026-01-03", "name": "CH. VARSHITHA", "class": "VI", "receipt_no": "5216", "payments": [["ONLINE (TRANSPORT)", 5000]], "sheet": "RCB Dec Jan 2025 26.pdf"},
    {"date": "2026-01-03", "name": "CH. JESHWITHA", "class": "III", "receipt_no": "5217", "payments": [["ONLINE (TRANSPORT)", 5000]], "sheet": "RCB Dec Jan 2025 26.pdf"},
    {"date": "2026-01-03", "name": "P. NITYA SRI", "class": "VI", "receipt_no": "5218", "payments": [["ONLINE (TRANSPORT)", 5000]], "sheet": "RCB Dec Jan 2025 26.pdf"}
]
page_6.extend(p6_transport)

# Rows 897-908
for r in range(897, 909):
    page_6.append(get_row_data(r))

for idx, tx in enumerate(page_6):
    tx["row"] = idx + 1

# --- PAGE 7 ---
# Rows 909-930 from Excel
page_7 = []
for r in range(909, 931):
    page_7.append(get_row_data(r))

for idx, tx in enumerate(page_7):
    tx["row"] = idx + 1

# --- PAGE 8 ---
page_8 = [
    {"date": "2026-01-27", "name": "J. NIKHIL", "class": "2ND", "receipt_no": "5282", "payments": [["CASH", 6000]], "sheet": "RCB Dec Jan 2025 26.pdf"},
    {"date": "2026-01-28", "name": "G. SAI PUJITHA", "class": "7TH", "receipt_no": "5283", "payments": [["ONLINE", 10000]], "sheet": "RCB Dec Jan 2025 26.pdf"},
    {"date": "2026-01-28", "name": "G. ANUSH", "class": "6TH", "receipt_no": "5284", "payments": [["ONLINE", 5000]], "sheet": "RCB Dec Jan 2025 26.pdf"},
    {"date": "2026-01-29", "name": "T. NEHANVI", "class": "NUR", "receipt_no": "5285", "payments": [["ONLINE", 5000]], "sheet": "RCB Dec Jan 2025 26.pdf"},
    {"date": "2026-01-28", "name": "R. ABHIRAM", "class": "4TH", "receipt_no": "5286", "payments": [["ONLINE", 20000]], "sheet": "RCB Dec Jan 2025 26.pdf"},
    {"date": "2026-01-28", "name": "R. ABHIRAM", "class": "4TH", "receipt_no": "", "payments": [["ONLINE (T.F)", 4000]], "sheet": "RCB Dec Jan 2025 26.pdf"},
    {"date": "2026-01-28", "name": "S. VENELLA", "class": "8TH", "receipt_no": "5287", "payments": [["CASH", 20000]], "sheet": "RCB Dec Jan 2025 26.pdf"},
    {"date": "2026-01-28", "name": "S. NAVYA SREE", "class": "2ND", "receipt_no": "5288", "payments": [["CASH", 10000]], "sheet": "RCB Dec Jan 2025 26.pdf"},
    {"date": "2026-01-28", "name": "S. VYSHNAV DEEP", "class": "UKG", "receipt_no": "5289", "payments": [["CASH", 10000]], "sheet": "RCB Dec Jan 2025 26.pdf"},
    {"date": "2026-01-29", "name": "V. SAI SRINATH", "class": "2ND", "receipt_no": "5290", "payments": [["CASH", 8000]], "sheet": "RCB Dec Jan 2025 26.pdf"},
    {"date": "2026-01-29", "name": "V. NITHYA SRI", "class": "UKG", "receipt_no": "5291", "payments": [["CASH", 8000]], "sheet": "RCB Dec Jan 2025 26.pdf"},
    {"date": "2026-01-29", "name": "B. AAKARSH", "class": "I", "receipt_no": "5292", "payments": [["ONLINE", 10000]], "sheet": "RCB Dec Jan 2025 26.pdf"},
    {"date": "2026-01-29", "name": "B. AAKARSH", "class": "3RD", "receipt_no": "", "payments": [["ONLINE (T.F)", 2000]], "sheet": "RCB Dec Jan 2025 26.pdf"},
    {"date": "2026-01-29", "name": "B. AAKARSH", "class": "3RD", "receipt_no": "5293", "payments": [["ONLINE", 7000]], "sheet": "RCB Dec Jan 2025 26.pdf"},
    {"date": "2026-01-29", "name": "B. AAKARSH", "class": "3RD", "receipt_no": "", "payments": [["ONLINE (T.F)", 2000]], "sheet": "RCB Dec Jan 2025 26.pdf"},
    {"date": "2026-01-29", "name": "T. HARINI", "class": "2ND", "receipt_no": "5294", "payments": [["CASH", 5000]], "sheet": "RCB Dec Jan 2025 26.pdf"},
    {"date": "2026-01-29", "name": "T. MANOJ", "class": "5TH", "receipt_no": "5295", "payments": [["CASH", 5000]], "sheet": "RCB Dec Jan 2025 26.pdf"},
    {"date": "2026-01-31", "name": "K. BHANU TEJA", "class": "3RD", "receipt_no": "5300", "payments": [["CASH", 6000]], "sheet": "RCB Dec Jan 2025 26.pdf"},
    {"date": "2026-01-31", "name": "K. BHANU TEJA", "class": "3RD", "receipt_no": "", "payments": [["CASH (T.F)", 5000]], "sheet": "RCB Dec Jan 2025 26.pdf"},
    {"date": "2026-01-30", "name": "M. HAASYA GOWRI", "class": "5TH", "receipt_no": "5296", "payments": [["ONLINE", 11000]], "sheet": "RCB Dec Jan 2025 26.pdf"},
    {"date": "2026-01-30", "name": "M. HAASYA GOWRI", "class": "5TH", "receipt_no": "", "payments": [["ONLINE (T.F)", 4000]], "sheet": "RCB Dec Jan 2025 26.pdf"},
    {"date": "2026-01-31", "name": "P. HARSHA", "class": "2ND", "receipt_no": "5297", "payments": [["ONLINE", 14000]], "sheet": "RCB Dec Jan 2025 26.pdf"},
    {"date": "2026-01-31", "name": "P. VARSHA", "class": "2ND", "receipt_no": "5298", "payments": [["ONLINE", 14000]], "sheet": "RCB Dec Jan 2025 26.pdf"},
    {"date": "2026-01-31", "name": "P. HARSHA", "class": "2ND", "receipt_no": "5299", "payments": [["ONLINE (T.F)", 2500]], "sheet": "RCB Dec Jan 2025 26.pdf"},
    {"date": "2026-01-31", "name": "P. VARSHA", "class": "2ND", "receipt_no": "", "payments": [["ONLINE (T.F)", 2500]], "sheet": "RCB Dec Jan 2025 26.pdf"},
    {"date": "2026-01-29", "name": "CH. KOMALI", "class": "6TH", "receipt_no": "003", "payments": [["CASH (T.F)", 5000]], "sheet": "RCB Dec Jan 2025 26.pdf"},
    {"date": "2026-01-29", "name": "CH. SOUJANYA", "class": "3RD", "receipt_no": "004", "payments": [["CASH (T.F)", 5000]], "sheet": "RCB Dec Jan 2025 26.pdf"},
    {"date": "2026-01-29", "name": "M. SHRESTHA", "class": "3RD", "receipt_no": "005", "payments": [["CASH (T.F)", 5000]], "sheet": "RCB Dec Jan 2025 26.pdf"}
]

for idx, tx in enumerate(page_8):
    tx["row"] = idx + 1

# --- PAGE 9 ---
page_9 = [
    {"date": "2026-01-30", "name": "D. SAHASRA", "class": "6TH", "receipt_no": "006", "payments": [["CASH", 15000]], "sheet": "RCB Dec Jan 2025 26.pdf"},
    {"date": "2026-01-30", "name": "D. PRAVARTHIKA", "class": "3RD", "receipt_no": "007", "payments": [["CASH", 15000]], "sheet": "RCB Dec Jan 2025 26.pdf"}
]

for idx, tx in enumerate(page_9):
    tx["row"] = idx + 1

# Save all pages
all_pages = [
    ("RCB_Dec_Jan_2025_26_p0_i0.jpg.json", page_0),
    ("RCB_Dec_Jan_2025_26_p1_i0.jpg.json", page_1),
    ("RCB_Dec_Jan_2025_26_p2_i0.jpg.json", page_2),
    ("RCB_Dec_Jan_2025_26_p3_i0.jpg.json", page_3),
    ("RCB_Dec_Jan_2025_26_p4_i0.jpg.json", page_4),
    ("RCB_Dec_Jan_2025_26_p5_i0.jpg.json", page_5),
    ("RCB_Dec_Jan_2025_26_p6_i0.jpg.json", page_6),
    ("RCB_Dec_Jan_2025_26_p7_i0.jpg.json", page_7),
    ("RCB_Dec_Jan_2025_26_p8_i0.jpg.json", page_8),
    ("RCB_Dec_Jan_2025_26_p9_i0.jpg.json", page_9)
]

for filename, data in all_pages:
    filepath = os.path.join(output_dir, filename)
    with open(filepath, 'w', encoding='utf-8') as f:
        json.dump(data, f, indent=2, ensure_ascii=False)
    print(f"Generated {filename} with {len(data)} transactions.")

wb.close()
print("All files generated successfully.")
