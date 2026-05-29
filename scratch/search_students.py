import openpyxl
import sys

sys.stdout.reconfigure(encoding='utf-8')

wb = openpyxl.load_workbook(r"E:\accounts\sorted fee 2025 2026.xlsx", data_only=True)
sheet = wb['Feb_Mar_Apr_2026']

student_names = [
    "NIKHIL", "PUJITHA", "ANUSH", "NEHANVI", "ABHIRAM", 
    "VENELLA", "NAVYA SREE", "VYSHNAV", "SRINATH", "NITHYA SRI", 
    "AAKARSH", "HARINI", "MANOJ", "BHANU TEJA", "HAASYA", 
    "HARSHA", "VARSHA", "KOMALI", "SOUJANYA", "SHRESTHA"
]

print("Searching for student names from Page 8 in Feb_Mar_Apr_2026:")
for r in range(3, sheet.max_row + 1):
    name_val = sheet.cell(row=r, column=7).value
    if name_val is not None:
        name_upper = str(name_val).upper()
        for s_name in student_names:
            if s_name in name_upper:
                vals = [sheet.cell(row=r, column=c).value for c in range(1, 13)]
                print(f"Row {r:4d}: {vals}")
                break

wb.close()
