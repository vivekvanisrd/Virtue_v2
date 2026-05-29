import openpyxl
import sys

sys.stdout.reconfigure(encoding='utf-8')

wb = openpyxl.load_workbook(r"E:\accounts\sorted fee 2025 2026.xlsx", data_only=True)
sheet = wb['Feb_Mar_Apr_2026']

names_to_search = ["sahasra", "pravarthika", "komali", "soujanya", "shrestha", "nikhil", "pujitha", "anush", "nehanvi", "abhiram", "venella", "navya", "vyshnav", "srinath", "nithya", "aakarsh", "harini", "manoj", "bhanu teja", "haasya", "harsha", "varsha"]

print("Searching for any partial name match:")
for r in range(3, sheet.max_row + 1):
    name_val = sheet.cell(row=r, column=7).value
    if name_val is not None:
        name_lower = str(name_val).lower()
        for target in names_to_search:
            if target in name_lower:
                # print row and name
                vals = [sheet.cell(row=r, column=c).value for c in range(1, 13)]
                print(f"Row {r:4d}: {vals}")
                break

wb.close()
