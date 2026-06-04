import openpyxl
wb = openpyxl.load_workbook(r"E:\accounts\sorted fee 2025 2026.xlsx", data_only=True)
sheet = wb['Feb_Mar_Apr_2026']

for r in range(3, 37):
    rec = sheet.cell(row=r, column=6).value
    name = sheet.cell(row=r, column=7).value
    cash = sheet.cell(row=r, column=9).value
    online = sheet.cell(row=r, column=10).value
    remarks = sheet.cell(row=r, column=12).value
    print(f"Row {r:3d} | Rec: {rec} | Name: {name} | Cash: {cash} | Online: {online} | Remarks: {remarks}")
wb.close()
