import openpyxl
import psycopg2
import uuid
import datetime

db_url = 'postgresql://postgres.bmyhbgwyirvjeadpvwny:VivekeVani%40369@aws-1-ap-northeast-1.pooler.supabase.com:5432/postgres'
conn = psycopg2.connect(db_url)
conn.autocommit = True
cur = conn.cursor()

def safe_float(val):
    if val is None:
        return 0.0
    s = str(val).strip()
    if not s or s.startswith('#'):
        return 0.0
    try:
        return float(s)
    except ValueError:
        return 0.0

# 1. Resolve Scope
cur.execute('SELECT id FROM "School" WHERE code = \'VIVES\' LIMIT 1;')
school_row = cur.fetchone()
school_id = school_row[0] if school_row else 'VIVES'

cur.execute('SELECT id FROM "Branch" WHERE "schoolId" = %s AND code = \'RCB\' LIMIT 1;', (school_id,))
branch_row = cur.fetchone()
branch_id = branch_row[0] if branch_row else 'VIVES-RCB'

cur.execute('SELECT id FROM "FinancialYear" WHERE "schoolId" = %s AND "isCurrent" = true LIMIT 1;', (school_id,))
fy_row = cur.fetchone()
fy_id = fy_row[0] if fy_row else 'VIVES-HQ-FY-2026-27'

cur.execute('SELECT id FROM "AcademicYear" WHERE "schoolId" = %s AND "isCurrent" = true LIMIT 1;', (school_id,))
ay_row = cur.fetchone()
ay_id = ay_row[0] if ay_row else 'VIVES-HQ-AY-2026-27'

print(f"Scope Resolved -> School: {school_id}, Branch: {branch_id}, FY: {fy_id}, AY: {ay_id}", flush=True)

# Resolve Class Map
cur.execute('SELECT id, name FROM "Class" WHERE "schoolId" = %s;', (school_id,))
class_rows = cur.fetchall()
class_map = {}
for cid, cname in class_rows:
    cname_upper = cname.upper().strip()
    class_map[cname_upper] = cid
    if cname_upper == "PLAY GROUP":
        class_map["PLAY GROUP"] = cid
        class_map["PG"] = cid
    elif cname_upper == "NURSERY":
        class_map["NUR"] = cid
        class_map["NURSERY"] = cid
    elif cname_upper == "1ST GRADE":
        class_map["1ST"] = cid
        class_map["CLASS 1"] = cid
    elif cname_upper == "2ND GRADE":
        class_map["2ND"] = cid
        class_map["CLASS 2"] = cid
    elif cname_upper == "3RD GRADE":
        class_map["3RD"] = cid
        class_map["CLASS 3"] = cid
    elif cname_upper == "4TH GRADE":
        class_map["4TH"] = cid
        class_map["CLASS 4"] = cid
    elif cname_upper == "5TH GRADE":
        class_map["5TH"] = cid
        class_map["CLASS 5"] = cid

default_class_id = class_rows[0][0] if class_rows else None
print(f"Class Map initialized with {len(class_map)} mappings.", flush=True)

# Load existing students map from DB
cur.execute('SELECT id, "admissionNumber", "studentCode" FROM "Student" WHERE "schoolId" = %s;', (school_id,))
existing_st_rows = cur.fetchall()
student_id_by_admi = {}
for sid, adm_no, st_code in existing_st_rows:
    if adm_no:
        student_id_by_admi[adm_no.strip()] = sid
    if st_code:
        student_id_by_admi[st_code.strip()] = sid

print(f"Found {len(student_id_by_admi)} existing student entries in DB.", flush=True)

wb = openpyxl.load_workbook('scratch/imported_sheets.xlsx', data_only=True)

# ─── STEP 1: UPSERT STUDENTS FROM STUDENT_MASTER ───
ws_students = wb['STUDENT_MASTER']
st_rows = [r for r in ws_students.iter_rows(values_only=True) if any(r)]

imported_students_count = 0
now = datetime.datetime.now()

for row in st_rows[1:]:
    admi_no = str(row[0]).strip() if row[0] is not None else None
    student_name = str(row[1]).strip() if row[1] is not None else None
    
    if not admi_no or not student_name or admi_no.upper() == "ADMI NO":
        continue
    
    parent_name = str(row[2]).strip() if (row[2] is not None and str(row[2]).strip() != '') else None
    contact = str(row[3]).strip() if (row[3] is not None and str(row[3]).strip() != '') else None
    class_str = str(row[5]).strip().upper() if row[5] is not None else 'NUR'
    
    adm_fee = safe_float(row[7])
    concession = safe_float(row[8])
    tuition_fee = safe_float(row[9])
    transport_fee = safe_float(row[10])
    
    target_class_id = class_map.get(class_str, default_class_id)
    
    name_parts = student_name.split(' ', 1)
    first_name = name_parts[0]
    last_name = name_parts[1] if len(name_parts) > 1 else ''
    
    st_id = student_id_by_admi.get(admi_no)
    if not st_id:
        st_id = f"std_{uuid.uuid4().hex[:12]}"
        
    cur.execute('''
        INSERT INTO "Student" (
            id, "schoolId", "branchId", "admissionNumber", "studentCode", "firstName", "lastName", gender, status, "createdAt", "updatedAt"
        ) VALUES (%s, %s, %s, %s, %s, %s, %s, 'OTHER', 'ACTIVE', %s, %s)
        ON CONFLICT ("schoolId", "branchId", "studentCode") DO UPDATE
        SET "firstName" = EXCLUDED."firstName", "lastName" = EXCLUDED."lastName", "updatedAt" = EXCLUDED."updatedAt"
        RETURNING id;
    ''', (st_id, school_id, branch_id, admi_no, admi_no, first_name, last_name, now, now))
    
    res = cur.fetchone()
    if res:
        st_id = res[0]
        student_id_by_admi[admi_no] = st_id

    # Academic Record
    cur.execute('SELECT id FROM "AcademicRecord" WHERE "studentId" = %s LIMIT 1;', (st_id,))
    ar_row = cur.fetchone()
    if not ar_row:
        ar_id = f"ar_{uuid.uuid4().hex[:12]}"
        cur.execute('''
            INSERT INTO "AcademicRecord" (
                id, "studentId", "classId", "academicYear", "branchId", "schoolId", "rollNumber"
            ) VALUES (%s, %s, %s, '2026-27', %s, %s, %s);
        ''', (ar_id, st_id, target_class_id, branch_id, school_id, admi_no))

    # Family Details (Parent Name & Phone)
    cur.execute('DELETE FROM "FamilyDetail" WHERE "studentId" = %s;', (st_id,))
    fam_id = f"fam_{uuid.uuid4().hex[:12]}"
    cur.execute('''
        INSERT INTO "FamilyDetail" (id, "studentId", "fatherName", "fatherPhone", "schoolId", "branchId")
        VALUES (%s, %s, %s, %s, %s, %s);
    ''', (fam_id, st_id, parent_name, contact, school_id, branch_id))
            
    # Financial Record
    net_tuition = max(0, tuition_fee - concession)
    cur.execute('DELETE FROM "FinancialRecord" WHERE "studentId" = %s;', (st_id,))
    fin_id = f"fin_{uuid.uuid4().hex[:12]}"
    cur.execute('''
        INSERT INTO "FinancialRecord" (
            id, "studentId", "annualTuition", "totalDiscount", "netTuition", "paymentType",
            "term1Amount", "term2Amount", "term3Amount", "admissionFee", "transportFee", "tuitionFee", "schoolId", "branchId"
        ) VALUES (%s, %s, %s, %s, %s, 'Term-wise', %s, %s, %s, %s, %s, %s, %s, %s);
    ''', (fin_id, st_id, tuition_fee, concession, net_tuition, net_tuition * 0.5, net_tuition * 0.25, net_tuition * 0.25, adm_fee, transport_fee, tuition_fee, school_id, branch_id))

    imported_students_count += 1
    if imported_students_count % 50 == 0:
        print(f"Ingested {imported_students_count} students into DB...", flush=True)

print(f"SUCCESS: Ingested {imported_students_count} students into Database!", flush=True)

# ─── STEP 2: IMPORT COLLECTIONS FROM FEE_COLLECTION ───
ws_fc = wb['FEE_COLLECTION']
fc_rows = [r for r in ws_fc.iter_rows(values_only=True) if any(r)]

imported_collections_count = 0

for row in fc_rows[2:]:
    if row[0] is None or row[3] is None:
        continue
    
    sl_no = row[0]
    date_val = row[1] if isinstance(row[1], (datetime.datetime, datetime.date)) else datetime.date.today()
    receipt_no = str(row[2]).split('.')[0] if row[2] is not None else None
    admi_no = str(row[3]).strip()
    student_name = str(row[4]).strip() if row[4] is not None else ''
    
    cash_amt = safe_float(row[5])
    online_amt = safe_float(row[6])
    total_amt = safe_float(row[7]) if (row[7] is not None and str(row[7]).strip() != '') else (cash_amt + online_amt)
    
    if total_amt <= 0:
        continue
        
    collected_by = str(row[10]).strip() if (row[10] is not None and str(row[10]).strip() != '') else 'Staff'
    txn_ref = str(row[11]).strip() if (row[11] is not None and str(row[11]).strip() != '') else None
    
    st_id = student_id_by_admi.get(admi_no)
    if not st_id:
        name_parts = student_name.split(' ', 1) if student_name else ['Unknown', 'Student']
        st_id = f"std_{uuid.uuid4().hex[:12]}"
        cur.execute('''
            INSERT INTO "Student" (
                id, "schoolId", "branchId", "admissionNumber", "studentCode", "firstName", "lastName", gender, status, "createdAt", "updatedAt"
            ) VALUES (%s, %s, %s, %s, %s, %s, %s, 'OTHER', 'ACTIVE', %s, %s)
            ON CONFLICT ("schoolId", "branchId", "studentCode") DO UPDATE
            SET "firstName" = EXCLUDED."firstName"
            RETURNING id;
        ''', (st_id, school_id, branch_id, admi_no, admi_no, name_parts[0], name_parts[1] if len(name_parts) > 1 else '', now, now))
        res = cur.fetchone()
        if res:
            st_id = res[0]
            student_id_by_admi[admi_no] = st_id

    pmode = "Cash" if cash_amt > 0 else ("Bank QR" if online_amt > 0 else "Cash")
    sys_receipt_no = f"VIRT-RCB-GS-{receipt_no or sl_no}"
    col_id = f"col_{uuid.uuid4().hex[:12]}"

    cur.execute('''
        INSERT INTO "Collection" (
            id, "receiptNumber", "bookReceiptNo", "studentId", "financialYearId", "schoolId", "branchId",
            "amountPaid", "lateFeePaid", "convenienceFee", "totalPaid", "paymentMode", "paymentReference",
            "collectedBy", "status", "paymentDate"
        ) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, 0, 0, %s, %s, %s, %s, 'Success', %s)
        ON CONFLICT DO NOTHING;
    ''', (col_id, sys_receipt_no, receipt_no, st_id, fy_id, school_id, branch_id, total_amt, total_amt, pmode, txn_ref, collected_by, date_val))

    imported_collections_count += 1
    if imported_collections_count % 50 == 0:
        print(f"Ingested {imported_collections_count} collections into DB...", flush=True)

conn.close()

print(f"\nSUCCESS: GOOGLE SHEETS IMPORT COMPLETED FOR ALL {imported_students_count} STUDENTS & {imported_collections_count} COLLECTIONS!", flush=True)
