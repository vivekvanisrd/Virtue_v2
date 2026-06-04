import openpyxl
import json
import os
import sys
import datetime

sys.stdout.reconfigure(encoding='utf-8')

excel_path = r"E:\accounts\RCB june 2025 26.xlsx"
output_dir = r"J:\virtue_fb\virtue-v2\scratch\transcribed_json"
os.makedirs(output_dir, exist_ok=True)

wb = openpyxl.load_workbook(excel_path, data_only=True)
sheet = wb['Sheet1']

def clean_name(name):
    if name is None:
        return ""
    name_str = str(name).strip().upper()
    while "  " in name_str:
        name_str = name_str.replace("  ", " ")
    return name_str

def clean_class(cls):
    if cls is None:
        return ""
    return str(cls).strip().upper()

def parse_date(date_val):
    if isinstance(date_val, datetime.datetime):
        return date_val.strftime('%Y-%m-%d')
    if date_val is None:
        return ""
    d_str = str(date_val).strip()
    for fmt in ('%d/%m/%Y', '%Y-%m-%d', '%d-%m-%Y', '%d/%m/%y', '%d/%m/%Y %H:%M:%S'):
        try:
            return datetime.datetime.strptime(d_str, fmt).strftime('%Y-%m-%d')
        except ValueError:
            pass
    return d_str

def make_payments(cash_val, online_val, remarks):
    payments = []
    cash = 0
    online = 0
    try:
        if cash_val:
            cash = int(float(str(cash_val).strip()))
    except ValueError:
        pass
    try:
        if online_val:
            online = int(float(str(online_val).strip()))
    except ValueError:
        pass
        
    remarks_lower = str(remarks).lower() if remarks else ""
    is_tf = "tf" in remarks_lower or "t/f" in remarks_lower or "term" in remarks_lower
    is_transport = "transport" in remarks_lower or "tp" in remarks_lower
    
    if cash > 0:
        if is_transport:
            payments.append(["CASH (TRANSPORT)", cash])
        elif is_tf:
            payments.append(["CASH (T.F)", cash])
        else:
            payments.append(["CASH", cash])
            
    if online > 0:
        if is_transport:
            payments.append(["ONLINE (TRANSPORT)", online])
        elif is_tf:
            payments.append(["ONLINE (T.F)", online])
        else:
            payments.append(["ONLINE", online])
            
    return payments

def get_row_data(row_idx):
    date_str = parse_date(sheet.cell(row=row_idx, column=2).value)
    rec = str(sheet.cell(row=row_idx, column=3).value or "").strip()
    if rec.endswith(".0"):
        rec = rec[:-2]
    name = clean_name(sheet.cell(row=row_idx, column=4).value)
    cls = clean_class(sheet.cell(row=row_idx, column=5).value)
    cash = sheet.cell(row=row_idx, column=6).value
    online = sheet.cell(row=row_idx, column=7).value
    remarks = sheet.cell(row=row_idx, column=9).value
    
    # Special handle for DEPOSIT
    if rec == "DEPOSIT" or name == "TO DEPOSIT":
        return {
            "date": date_str,
            "name": "TO DEPOSIT",
            "class": "UNCLEAR",
            "receipt_no": "DEPOSIT",
            "payments": [["DEPOSIT", cash or online or 0]],
            "sheet": "RCB june 2025 26.pdf"
        }
        
    # Handle side T/F entries
    if rec == "T/F" or "T/F" in name:
        # These are side entries or split payments
        # We'll treat them as a payment
        amt = cash or online or 0
        pmt_type = "CASH (T.F)" if cash else "ONLINE (T.F)"
        return {
            "date": date_str,
            "name": name,
            "class": cls,
            "receipt_no": "T/F",
            "payments": [[pmt_type, amt]],
            "sheet": "RCB june 2025 26.pdf"
        }

    payments = make_payments(cash, online, remarks)
    
    return {
        "date": date_str,
        "name": name,
        "class": cls,
        "receipt_no": rec,
        "payments": payments,
        "sheet": "RCB june 2025 26.pdf"
    }

# Page Ranges:
# Page 0: Rows 2 to 26 (Row index in Excel is 2 to 26)
# Page 1: Rows 27 to 51
# Page 2: Rows 53 to 82 (Row 52 is empty)
# Page 3: Rows 83 to 92 (Note: row 87 is deposit)

page_0 = [get_row_data(r) for r in range(2, 27)]
page_1 = [get_row_data(r) for r in range(27, 52)]
page_2 = [get_row_data(r) for r in range(53, 83)]
page_3 = [get_row_data(r) for r in range(83, 93)]

# Assign row numbers page-by-page
for idx, tx in enumerate(page_0): tx["row"] = idx + 1
for idx, tx in enumerate(page_1): tx["row"] = idx + 1
for idx, tx in enumerate(page_2): tx["row"] = idx + 1
for idx, tx in enumerate(page_3): tx["row"] = idx + 1

# Verification sums helper
def verify_page_sums(page_data, expected_cash, expected_online, expected_deposit):
    cash_sum = 0
    online_sum = 0
    deposit_sum = 0
    for tx in page_data:
        for ptype, amt in tx["payments"]:
            if ptype == "DEPOSIT":
                deposit_sum += amt
            elif "CASH" in ptype:
                cash_sum += amt
            elif "ONLINE" in ptype:
                online_sum += amt
    print(f"Computed Sums: Cash={cash_sum}, Online={online_sum}, Deposit={deposit_sum}")
    print(f"Expected Sums: Cash={expected_cash}, Online={expected_online}, Deposit={expected_deposit}")
    print(f"Status: Cash={cash_sum == expected_cash}, Online={online_sum == expected_online}, Deposit={deposit_sum == expected_deposit}")

print("\n--- JUNE PAGE 0 VERIFICATION ---")
verify_page_sums(page_0, 105500, 179500, 50000)

print("\n--- JUNE PAGE 1 VERIFICATION ---")
verify_page_sums(page_1, 200000, 243000, 214500)

print("\n--- JUNE PAGE 2 VERIFICATION ---")
verify_page_sums(page_2, 158000, 145500, 89500)

print("\n--- JUNE PAGE 3 VERIFICATION ---")
# On Page 3, Saketh has cash 18000, online 12000. Let's make sure it matches.
verify_page_sums(page_3, 120000, 34000, 160000)

# Save June page JSONs
all_pages = [
    ("RCB_june_2025_26_p0_i0.jpg.json", page_0),
    ("RCB_june_2025_26_p1_i0.jpg.json", page_1),
    ("RCB_june_2025_26_p2_i0.jpg.json", page_2),
    ("RCB_june_2025_26_p3_i0.jpg.json", page_3)
]

for filename, data in all_pages:
    filepath = os.path.join(output_dir, filename)
    with open(filepath, 'w', encoding='utf-8') as f:
        json.dump(data, f, indent=2, ensure_ascii=False)
    print(f"Generated {filename} with {len(data)} transactions.")

wb.close()
