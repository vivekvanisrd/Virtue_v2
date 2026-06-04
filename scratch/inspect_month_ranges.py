import json
import openpyxl
from collections import Counter
from datetime import datetime

# Load student_data.json
JSON_PATH = "public/student-directory/student_data.json"
with open(JSON_PATH, "r", encoding="utf-8") as f:
    student_profiles = json.load(f)

json_months = Counter()
for profile in student_profiles:
    for tr in profile.get("payments", []):
        date_str = tr.get("date", "")
        if date_str:
            # Extract YYYY-MM
            ym = date_str[:7]
            json_months[ym] += 1

print("=== Date ranges in student_data.json ===")
for ym, count in sorted(json_months.items()):
    print(f"Month {ym}: {count} records")

# Load Excel workbook
EXCEL_PATH = r"E:\accounts\sorted fee 2025 2026.xlsx"
wb = openpyxl.load_workbook(EXCEL_PATH, read_only=True)
sheet = wb["Ucban"]

excel_months = Counter()
for r in range(4, sheet.max_row + 1):
    date_val = sheet.cell(row=r, column=6).value
    if date_val is None:
        continue
    if isinstance(date_val, datetime):
        ym = date_val.strftime("%Y-%m")
        excel_months[ym] += 1
    else:
        # String representation
        val_str = str(date_val).strip()
        # check format e.g. 02/02/2026 or 2025-08-02
        if "-" in val_str:
            ym = val_str[:7]
            excel_months[ym] += 1
        elif "/" in val_str:
            parts = val_str.split("/")
            if len(parts) == 3:
                # Assuming dd/mm/yyyy
                ym = f"{parts[2]}-{parts[1]}"
                excel_months[ym] += 1

print("\n=== Date ranges in Excel sheet 'Ucban' ===")
for ym, count in sorted(excel_months.items()):
    print(f"Month {ym}: {count} records")
