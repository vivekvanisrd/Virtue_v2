import openpyxl
import sys

sys.stdout.reconfigure(encoding='utf-8')

wb = openpyxl.load_workbook(r"E:\accounts\sorted fee 2025 2026.xlsx", data_only=True)
sheet = wb['Feb_Mar_Apr_2026']

print("Searching for receipts 3 to 7:")
for r in range(3, sheet.max_row + 1):
    rec_val = sheet.cell(row=r, column=6).value
    if rec_val is not None:
        rec_str = str(rec_val).strip()
        if rec_str in ['3', '4', '5', '6', '7', '003', '004', '005', '006', '007', '03', '04', '05', '06', '07']:
            vals = [sheet.cell(row=r, column=c).value for c in range(1, 13)]
            print(f"Row {r:4d}: {vals}")

wb.close()
