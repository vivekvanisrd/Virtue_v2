import openpyxl

wb = openpyxl.load_workbook(r"E:\accounts\sorted fee 2025 2026.xlsx", data_only=True)
sheet = wb['Feb_Mar_Apr_2026']

cash_total = 0
online_total = 0
for r in range(870, 931):
    cash = sheet.cell(row=r, column=9).value or 0
    online = sheet.cell(row=r, column=10).value or 0
    cash_total += cash
    online_total += online

print(f"Excel January 2026 rows (870-930) totals: Cash = {cash_total}, Online = {online_total}, Total = {cash_total + online_total}")
wb.close()
