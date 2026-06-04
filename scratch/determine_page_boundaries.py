import os
import openpyxl
import sys

sys.stdout.reconfigure(encoding='utf-8')
dir_path = r"E:\accounts"

files_and_pages = [
    ("RCB june 2025 26.xlsx", "Sheet1", 4),
    ("RCB july 2025 26.xlsx", "Sheet1", 14),
    ("August_2025_Fee_Extraction.xlsx", "August_2025", 9),
    ("RCB Sep 2025 26.xlsx", "Sheet1", 9),
    ("RCB Oct 2025 26.2pdf.xlsx", "Sheet1", 10),
    ("November_2025_Complete_Extraction.xlsx", "November_2025", 4)
]

for filename, sheet_name, expected_pages in files_and_pages:
    file_path = os.path.join(dir_path, filename)
    wb = openpyxl.load_workbook(file_path, data_only=True)
    sheet = wb[sheet_name]
    
    print(f"\n================ {filename} (expected {expected_pages} pages) ================")
    
    # Let's find all rows with DEPOSIT, or empty, or total/summary
    dividers = []
    for r in range(1, sheet.max_row + 1):
        vals = [sheet.cell(row=r, column=c).value for c in range(1, min(10, sheet.max_column+1))]
        is_empty = all(x is None or str(x).strip() == "" for x in vals)
        rec = str(sheet.cell(row=r, column=3).value or "").strip()
        name = str(sheet.cell(row=r, column=4).value or "").strip()
        
        if is_empty or "DEPOSIT" in rec.upper() or "TOTAL" in name.upper() or "TOTAL" in rec.upper() or "NOTE" in rec.upper():
            dividers.append((r, rec, name, vals[5], vals[6], vals[7])) # row, rec, name, cash, online, total
            
    print(f"Found {len(dividers)} divider rows:")
    for div in dividers:
        print(f"  Row {div[0]:3d} | Rec: {div[1]:10s} | Name: {div[2]:20s} | Cash: {div[3]} | Online: {div[4]} | Total: {div[5]}")
        
    wb.close()
