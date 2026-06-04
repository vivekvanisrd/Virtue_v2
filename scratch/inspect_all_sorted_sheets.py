import json
import openpyxl
import re
from datetime import datetime

# Load student_data.json
JSON_PATH = r"public/student-directory/student_data.json"
with open(JSON_PATH, "r", encoding="utf-8") as f:
    student_profiles = json.load(f)

# Extract all payments from student_data.json
json_payments = []
for profile in student_profiles:
    for tr in profile.get("payments", []):
        date = tr.get("date", "")
        receipt_no = tr.get("receipt_no", "")
        sheet = tr.get("sheet", "")
        row_num = tr.get("row", "")
        
        # Calculate cash and online amounts for this transaction
        cash_amt = 0.0
        online_amt = 0.0
        for label, val in tr.get("payments", []):
            try:
                amt = float(val)
            except (ValueError, TypeError):
                continue
            
            lbl_upper = str(label).upper()
            if "CASH" in lbl_upper:
                cash_amt += amt
            elif "ONLINE" in lbl_upper or "TRANSFER" in lbl_upper or "G PAY" in lbl_upper or "PHONEPE" in lbl_upper or "PAYTM" in lbl_upper:
                online_amt += amt
            else:
                # Default to online if not cash, or split based on label?
                # Let's check how the label is named
                online_amt += amt
                
        json_payments.append({
            "student_name": profile.get("name", ""),
            "name_in_sheet": tr.get("name", ""),
            "class": profile.get("class", ""),
            "class_in_sheet": tr.get("class", ""),
            "branch": profile.get("branch", ""),
            "date": date,
            "receipt_no": receipt_no,
            "cash": cash_amt,
            "online": online_amt,
            "total": cash_amt + online_amt,
            "sheet": sheet,
            "row": row_num
        })

print(f"Extracted {len(json_payments)} individual payments from student_data.json.")

# Load Excel workbook
EXCEL_PATH = r"E:\accounts\sorted fee 2025 2026.xlsx"
wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True)

def normalize_name(name):
    if not name:
        return ""
    # Remove dots, spaces, convert to uppercase
    return re.sub(r"[\s\.]", "", str(name)).upper()

def normalize_receipt(rec):
    if rec is None:
        return ""
    rec_str = str(rec).strip()
    if rec_str.endswith(".0"):
        rec_str = rec_str[:-2]
    # Remove leading zeros if numerical
    rec_str = re.sub(r"\s+", "", rec_str)
    if rec_str.isdigit():
        return str(int(rec_str))
    return rec_str.upper()

def parse_date(val):
    if val is None:
        return ""
    if isinstance(val, datetime):
        return val.strftime("%Y-%m-%d")
    val_str = str(val).strip()
    # Try parsing different date formats if needed
    for fmt in ("%d/%m/%Y", "%Y-%m-%d", "%d-%m-%Y"):
        try:
            return datetime.strptime(val_str, fmt).strftime("%Y-%m-%d")
        except ValueError:
            pass
    # Return split string if datetime format representation like '2025-08-02 00:00:00'
    return val_str.split()[0]

# Let's extract payments from the 'Ucban' sheet
excel_payments = []
sheet = wb["Ucban"]

# Row 3 contains headers: FULL SLNO, Sl No, Branch, Month, Date, Receipt No, Student Name, Class, Cash, Online, Total
for r in range(4, sheet.max_row + 1):
    name = sheet.cell(row=r, column=8).value
    if not name:
        continue
    
    # Skip deposit or expenditure entries if they got listed
    name_str = str(name).upper()
    if any(x in name_str for x in ["DEPOSIT", "EXPENDITURE", "TOTAL", "GRAND TOTAL", "CONTRA", "CEMENT", "DIESEL"]):
        continue
        
    date_val = sheet.cell(row=r, column=6).value
    receipt_val = sheet.cell(row=r, column=7).value
    
    cash_val = sheet.cell(row=r, column=10).value
    online_val = sheet.cell(row=r, column=11).value
    total_val = sheet.cell(row=r, column=12).value
    
    cash = float(cash_val) if cash_val is not None else 0.0
    online = float(online_val) if online_val is not None else 0.0
    total = float(total_val) if total_val is not None else (cash + online)
    
    excel_payments.append({
        "row": r,
        "student_name": str(name).strip(),
        "class": str(sheet.cell(row=r, column=9).value).strip() if sheet.cell(row=r, column=9).value else "",
        "date": parse_date(date_val),
        "receipt_no": normalize_receipt(receipt_val),
        "cash": cash,
        "online": online,
        "total": total,
        "branch": sheet.cell(row=r, column=4).value or ""
    })

print(f"Extracted {len(excel_payments)} payments from Excel sheet 'Ucban'.")

# Let's perform a matching logic
# We'll try to match each excel payment with a json payment
matched_json_indices = set()
unmatched_excel = []
unmatched_json = []

for ep in excel_payments:
    norm_name = normalize_name(ep["student_name"])
    norm_receipt = normalize_receipt(ep["receipt_no"])
    ep_date = ep["date"]
    
    found_match = False
    
    # 1st pass: Match on exact normalized name, date, receipt number, and total
    for idx, jp in enumerate(json_payments):
        if idx in matched_json_indices:
            continue
            
        name_match = normalize_name(jp["name_in_sheet"]) == norm_name or normalize_name(jp["student_name"]) == norm_name
        receipt_match = normalize_receipt(jp["receipt_no"]) == norm_receipt if norm_receipt else True
        date_match = jp["date"] == ep_date if ep_date and jp["date"] else True
        amount_match = abs(jp["total"] - ep["total"]) < 0.01
        
        # If receipt_no is empty in both or matches, and we have name, amount and date matches
        if name_match and receipt_match and date_match and amount_match:
            matched_json_indices.add(idx)
            found_match = True
            break
            
    if not found_match:
        # 2nd pass: Looser match (name, amount, either date or receipt match)
        for idx, jp in enumerate(json_payments):
            if idx in matched_json_indices:
                continue
                
            name_match = normalize_name(jp["name_in_sheet"]) == norm_name or normalize_name(jp["student_name"]) == norm_name
            receipt_match = normalize_receipt(jp["receipt_no"]) == norm_receipt
            date_match = jp["date"] == ep_date
            amount_match = abs(jp["total"] - ep["total"]) < 0.01
            
            # If name matches and amount matches, and either date or receipt matches
            if name_match and amount_match and (date_match or receipt_match):
                matched_json_indices.add(idx)
                found_match = True
                break
                
    if not found_match:
        unmatched_excel.append(ep)

# Identify unmatched JSON payments
for idx, jp in enumerate(json_payments):
    if idx not in matched_json_indices:
        unmatched_json.append(jp)

print(f"Matched {len(matched_json_indices)} transactions.")
print(f"Unmatched in Excel (sorted fee): {len(unmatched_excel)}")
print(f"Unmatched in Dashboard (student_data.json): {len(unmatched_json)}")

# Output sample differences
print("\n--- SAMPLE DIFFERENCES: ONLY IN EXCEL (SORTED FEE) ---")
for i, ep in enumerate(unmatched_excel[:20]):
    print(f"Excel Row {ep['row']}: Name: {ep['student_name']}, Class: {ep['class']}, Date: {ep['date']}, Receipt: {ep['receipt_no']}, Total: {ep['total']} (Cash: {ep['cash']}, Online: {ep['online']})")

print("\n--- SAMPLE DIFFERENCES: ONLY IN DASHBOARD (JSON) ---")
for i, jp in enumerate(unmatched_json[:20]):
    print(f"JSON from {jp['sheet']} Row {jp['row']}: Name in sheet: {jp['name_in_sheet']} (Profile: {jp['student_name']}), Class: {jp['class_in_sheet']}, Date: {jp['date']}, Receipt: {jp['receipt_no']}, Total: {jp['total']}")
