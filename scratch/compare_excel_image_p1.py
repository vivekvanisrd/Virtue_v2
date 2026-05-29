import openpyxl

wb = openpyxl.load_workbook(r"E:\accounts\sorted fee 2025 2026.xlsx", data_only=True)
sheet = wb['Feb_Mar_Apr_2026']

# Page 1 matches rows 791 to 814
print("Excel values for Page 1 (Rows 791-814):")
total_cash = 0
total_online = 0
for r in range(791, 815):
    rec = sheet.cell(row=r, column=6).value
    name = sheet.cell(row=r, column=7).value
    cash = sheet.cell(row=r, column=9).value or 0
    online = sheet.cell(row=r, column=10).value or 0
    total_cash += cash
    total_online += online
    print(f"Row {r:3d} | Rec {rec} | {name} | Cash {cash} | Online {online}")

print(f"\nExcel totals for Page 1: Cash = {total_cash}, Online = {total_online}")
wb.close()
