import openpyxl
import sys

sys.stdout.reconfigure(encoding='utf-8')
wb = openpyxl.load_workbook(r"E:\accounts\RCB june 2025 26.xlsx", data_only=True)
sheet = wb['Sheet1']

print("All June rows:")
for r in range(1, sheet.max_row + 1):
    vals = [sheet.cell(row=r, column=c).value for c in range(1, 10)]
    # Check if this row is empty or has a deposit or is a header
    is_empty = all(x is None or str(x).strip() == "" for x in vals)
    rec = str(vals[2]).strip() if vals[2] is not None else ""
    name = str(vals[3]).strip() if vals[3] is not None else ""
    
    if is_empty or "DEPOSIT" in rec or "TOTAL" in name.upper() or r == 1 or r == sheet.max_row:
        print(f"Row {r:3d} | Vals: {vals}")
wb.close()
