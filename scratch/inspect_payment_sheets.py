import openpyxl

file_path = r"C:\Users\SriKriations\Favorites\Downloads\DAILY ACCOUNTS.xlsx"
wb = openpyxl.load_workbook(file_path, data_only=True)

for sheet_name in ["RCB", "SNB", "MNB"]:
    if sheet_name not in wb.sheetnames:
        continue
    sheet = wb[sheet_name]
    print(f"\n======================================")
    print(f"SHEET: {sheet_name}")
    print(f"======================================")
    
    for r in range(1, 15):
        row_vals = [sheet.cell(row=r, column=c).value for c in range(1, 12)]
        if any(v is not None for v in row_vals):
            aligned = []
            for idx, val in enumerate(row_vals):
                aligned.append(f"{idx+1}:{val}")
            print(f"Row {r:2d}: " + " | ".join(aligned))
