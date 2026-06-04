import os
import openpyxl
import sys

sys.stdout.reconfigure(encoding='utf-8')
dir_path = r"E:\accounts"

def inspect_file(filename, sheet_name, rec_col, name_col, date_col):
    file_path = os.path.join(dir_path, filename)
    wb = openpyxl.load_workbook(file_path, data_only=True)
    sheet = wb[sheet_name]
    print(f"\n===== {filename} / {sheet_name} =====")
    
    rows_info = []
    for r in range(2, sheet.max_row + 1):
        rec = str(sheet.cell(row=r, column=rec_col).value or "").strip()
        if rec.endswith(".0"):
            rec = rec[:-2]
        name = str(sheet.cell(row=r, column=name_col).value or "").strip()
        date = str(sheet.cell(row=r, column=date_col).value or "").strip()
        cash = sheet.cell(row=r, column=rec_col + 3).value if rec_col + 3 <= sheet.max_column else None # approximate
        # Just grab the actual receipt and name
        rows_info.append((r, date, rec, name))
        
    # Print some info: total rows, and every 10th row, and rows where receipt is DEPOSIT or empty
    print(f"Total rows: {len(rows_info)}")
    for idx, (r, date, rec, name) in enumerate(rows_info):
        # Print first 5, last 5, and any rows with DEPOSIT or weird receipt
        if idx < 10 or idx > len(rows_info) - 10 or "DEPOSIT" in rec or not rec:
            print(f"  Row {r:3d} | Date: {date[:10]} | Rec: {rec:10s} | Name: {name}")
    wb.close()

inspect_file("RCB june 2025 26.xlsx", "Sheet1", 3, 4, 2)
inspect_file("RCB july 2025 26.xlsx", "Sheet1", 3, 4, 2)
inspect_file("August_2025_Fee_Extraction.xlsx", "August_2025", 3, 4, 2)
inspect_file("RCB Sep 2025 26.xlsx", "Sheet1", 3, 4, 2)
inspect_file("RCB Oct 2025 26.2pdf.xlsx", "Sheet1", 3, 4, 2)
inspect_file("November_2025_Complete_Extraction.xlsx", "November_2025", 3, 4, 2)
