import openpyxl
import os
import sys

sys.stdout.reconfigure(encoding='utf-8')

dir_path = r"E:\accounts"
search_items = ["5282", "5283", "nikhil", "pujitha", "sahasra", "pravarthika"]

for filename in os.listdir(dir_path):
    if filename.endswith(".xlsx") and not filename.startswith("~$"):
        file_path = os.path.join(dir_path, filename)
        try:
            wb = openpyxl.load_workbook(file_path, data_only=True)
            for sheet_name in wb.sheetnames:
                sheet = wb[sheet_name]
                for r in range(1, sheet.max_row + 1):
                    for c in range(1, sheet.max_column + 1):
                        val = sheet.cell(row=r, column=c).value
                        if val is not None:
                            val_str = str(val).lower()
                            for item in search_items:
                                if item in val_str:
                                    print(f"File: {filename} | Sheet: {sheet_name} | Row {r}, Col {c}: {val} | Full: {[sheet.cell(row=r, column=col).value for col in range(1, min(10, sheet.max_column+1))]}")
            wb.close()
        except Exception as e:
            print(f"Error reading {filename}: {e}")
