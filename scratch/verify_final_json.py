import json
import openpyxl
import re
from datetime import datetime

# Load student_data.json
JSON_PATH = "public/student-directory/student_data.json"
with open(JSON_PATH, "r", encoding="utf-8") as f:
    student_profiles = json.load(f)

# Extract only PDF-based payments from student_data.json
json_payments = {}
for profile in student_profiles:
    for tr in profile.get("payments", []):
        sheet = tr.get("sheet", "")
        # Filter to only PDF registers
        if not sheet.lower().endswith(".pdf"):
            continue
            
        row_num = tr.get("row", "")
        date = tr.get("date", "")
        receipt_no = tr.get("receipt_no", "")
        name_in_sheet = tr.get("name", "")
        
        # Calculate cash and online
        cash_amt = 0.0
        online_amt = 0.0
        pay_details = []
        for label, val in tr.get("payments", []):
            try:
                amt = float(val)
            except (ValueError, TypeError):
                continue
            pay_details.append((label, amt))
            lbl_upper = str(label).upper()
            if "CASH" in lbl_upper:
                cash_amt += amt
            else:
                online_amt += amt
                
        # Truly unique transaction key
        tx_key = (sheet, date, name_in_sheet, receipt_no, str(sorted(pay_details)), row_num)
        json_payments[tx_key] = {
            "student_name": profile["name"],
            "name_in_sheet": name_in_sheet,
            "class": profile["class"],
            "class_in_sheet": tr.get("class", ""),
            "branch": profile["branch"],
            "date": date,
            "receipt_no": receipt_no,
            "cash": cash_amt,
            "online": online_amt,
            "total": cash_amt + online_amt,
            "sheet": sheet,
            "row": row_num
        }

json_list = list(json_payments.values())
print(f"Extracted {len(json_list)} unique PDF payments from student_data.json.")

# Load Excel workbook
EXCEL_PATH = r"E:\accounts\sorted fee 2025 2026.xlsx"
wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True)
sheet = wb["Ucban"]

def normalize_name(name):
    if not name:
        return ""
    return re.sub(r"[\s\.]", "", str(name)).upper()

def normalize_receipt(rec):
    if rec is None:
        return ""
    rec_str = str(rec).strip()
    if rec_str.endswith(".0"):
        rec_str = rec_str[:-2]
    rec_str = re.sub(r"\s+", "", rec_str)
    if rec_str.isdigit():
        return str(int(rec_str))
    return rec_str.upper()

def parse_date(val):
    if val is None:
        return ""
    if isinstance(val, datetime):
        return val.strftime("%Y-%m-%d")
    val_str = str(val).strip()
    for fmt in ("%d/%m/%Y", "%Y-%m-%d", "%d-%m-%Y"):
        try:
            return datetime.strptime(val_str, fmt).strftime("%Y-%m-%d")
        except ValueError:
            pass
    return val_str.split()[0]

excel_payments = []
for r in range(4, sheet.max_row + 1):
    name = sheet.cell(row=r, column=8).value
    if not name:
        continue
    
    name_str = str(name).upper()
    if any(x in name_str for x in ["DEPOSIT", "EXPENDITURE", "TOTAL", "GRAND TOTAL", "CONTRA", "CEMENT", "DIESEL"]):
        continue
        
    date_val = sheet.cell(row=r, column=6).value
    receipt_val = sheet.cell(row=r, column=7).value
    
    cash_val = sheet.cell(row=r, column=10).value
    online_val = sheet.cell(row=r, column=11).value
    total_val = sheet.cell(row=r, column=12).value
    
    cash = float(cash_val) if cash_val is not None else 0.0
    online = float(online_val) if online_val is not None else 0.0
    total = float(total_val) if total_val is not None else (cash + online)
    
    excel_payments.append({
        "row": r,
        "student_name": str(name).strip(),
        "class": str(sheet.cell(row=r, column=9).value).strip() if sheet.cell(row=r, column=9).value else "",
        "date": parse_date(date_val),
        "receipt_no": normalize_receipt(receipt_val),
        "cash": cash,
        "online": online,
        "total": total,
        "branch": sheet.cell(row=r, column=4).value or ""
    })

print(f"Extracted {len(excel_payments)} payments from Excel sheet 'Ucban'.")

# Perform matching
matched_json_indices = set()
unmatched_excel = []
unmatched_json = []

for ep in excel_payments:
    norm_name = normalize_name(ep["student_name"])
    norm_receipt = normalize_receipt(ep["receipt_no"])
    ep_date = ep["date"]
    
    found_match = False
    
    # 1. Match on Name + Receipt + Date + Total
    for idx, jp in enumerate(json_list):
        if idx in matched_json_indices:
            continue
            
        name_match = normalize_name(jp["name_in_sheet"]) == norm_name or normalize_name(jp["student_name"]) == norm_name
        receipt_match = normalize_receipt(jp["receipt_no"]) == norm_receipt if norm_receipt else True
        date_match = jp["date"] == ep_date if ep_date and jp["date"] else True
        amount_match = abs(jp["total"] - ep["total"]) < 0.01
        
        if name_match and receipt_match and date_match and amount_match:
            matched_json_indices.add(idx)
            found_match = True
            break
            
    if not found_match:
        # 2. Match on Name + Amount + (Date or Receipt)
        for idx, jp in enumerate(json_list):
            if idx in matched_json_indices:
                continue
                
            name_match = normalize_name(jp["name_in_sheet"]) == norm_name or normalize_name(jp["student_name"]) == norm_name
            receipt_match = normalize_receipt(jp["receipt_no"]) == norm_receipt
            date_match = jp["date"] == ep_date
            amount_match = abs(jp["total"] - ep["total"]) < 0.01
            
            if name_match and amount_match and (date_match or receipt_match):
                matched_json_indices.add(idx)
                found_match = True
                break
                
    if not found_match:
        unmatched_excel.append(ep)

for idx, jp in enumerate(json_list):
    if idx not in matched_json_indices:
        unmatched_json.append(jp)

print(f"Matched {len(matched_json_indices)} transactions.")
print(f"Unmatched in Excel (sorted fee): {len(unmatched_excel)}")
print(f"Unmatched in Dashboard (student_data.json): {len(unmatched_json)}")

# Save detailed differences
differences = {
    "only_in_excel": unmatched_excel,
    "only_in_dashboard": unmatched_json
}
with open("scratch/pdf_vs_excel_differences.json", "w", encoding="utf-8") as f:
    json.dump(differences, f, indent=2)

print("\nSaved detailed differences to scratch/pdf_vs_excel_differences.json")
