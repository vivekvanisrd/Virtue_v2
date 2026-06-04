import openpyxl
wb = openpyxl.load_workbook(r"E:\accounts\August_2025_Fee_Extraction.xlsx", data_only=True)
sheet = wb['August_2025']

for r in range(1, sheet.max_row + 1):
    vals = [sheet.cell(row=r, column=c).value for c in range(1, 10)]
    print(f"Row {r:2d}: {vals}")
wb.close()
