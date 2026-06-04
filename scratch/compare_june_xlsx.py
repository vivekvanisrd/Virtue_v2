import openpyxl

wb = openpyxl.load_workbook(r"E:\accounts\sorted fee 2025 2026.xlsx", data_only=True)

ucban = wb["Ucban"]
feb = wb["Feb_Mar_Apr_2026"]

print(f"Ucban rows: {ucban.max_row}, Feb_Mar_Apr_2026 rows: {feb.max_row}")

# Check first 5 data rows
print("Ucban first 3 rows of data:")
for r in range(4, 7):
    row_vals = [ucban.cell(row=r, column=c).value for c in range(2, 14)]
    print(f"Row {r}: {row_vals}")

print("\nFeb_Mar_Apr_2026 first 3 rows of data:")
for r in range(3, 6):
    row_vals = [feb.cell(row=r, column=c).value for c in range(1, 13)]
    print(f"Row {r}: {row_vals}")

# Let's count rows that have a Student Name in each
def get_non_empty_rows(sheet, start_row, name_col):
    data = []
    for r in range(start_row, sheet.max_row + 1):
        name = sheet.cell(row=r, column=name_col).value
        if name:
            row_vals = [sheet.cell(row=r, column=c).value for c in range(1, sheet.max_column + 1)]
            data.append((r, row_vals))
    return data

ucban_data = get_non_empty_rows(ucban, 4, 8)
feb_data = get_non_empty_rows(feb, 3, 7)

print(f"\nNon-empty row count - Ucban: {len(ucban_data)}, Feb: {len(feb_data)}")

# Compare rows by comparing a normalized representation (name, date, receipt_no, total)
# Let's see if they have the same names and totals
ucban_set = set()
for r, row in ucban_data:
    name = str(row[7]).strip().upper() if row[7] else ""
    date = str(row[5]).strip() if row[5] else ""
    receipt = str(row[6]).strip() if row[6] else ""
    total = row[11] if len(row) > 11 else None
    ucban_set.add((name, date, receipt, total))

feb_set = set()
for r, row in feb_data:
    name = str(row[6]).strip().upper() if row[6] else ""
    date = str(row[4]).strip() if row[4] else ""
    receipt = str(row[5]).strip() if row[5] else ""
    total = row[10] if len(row) > 10 else None
    feb_set.add((name, date, receipt, total))

print(f"Unique keys in Ucban: {len(ucban_set)}")
print(f"Unique keys in Feb: {len(feb_set)}")
print(f"Intersection: {len(ucban_set.intersection(feb_set))}")
print(f"Only in Ucban: {len(ucban_set - feb_set)}")
print(f"Only in Feb: {len(feb_set - ucban_set)}")
