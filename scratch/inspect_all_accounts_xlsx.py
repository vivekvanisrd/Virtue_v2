import os
import openpyxl
import sys

sys.stdout.reconfigure(encoding='utf-8')
dir_path = r"E:\accounts"

for filename in sorted(os.listdir(dir_path)):
    if filename.endswith(".xlsx") and not filename.startswith("~$") and filename != "sorted fee 2025 2026.xlsx" and filename != "fee 2025 2026.xlsx" and filename != "DAILY ACCOUNTS.xlsx":
        file_path = os.path.join(dir_path, filename)
        try:
            wb = openpyxl.load_workbook(file_path, data_only=True)
            print(f"\n================ FILE: {filename} ================")
            for name in wb.sheetnames:
                sheet = wb[name]
                print(f"  Sheet: {name} | Max Rows: {sheet.max_row} | Max Cols: {sheet.max_column}")
                # Print headers and first 2 data rows
                for r in range(1, min(4, sheet.max_row + 1)):
                    row_vals = [sheet.cell(row=r, column=c).value for c in range(1, min(10, sheet.max_column+1))]
                    print(f"    Row {r}: {row_vals}")
            wb.close()
        except Exception as e:
            print(f"Error reading {filename}: {e}")
