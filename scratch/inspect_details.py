import openpyxl

file_path = r"C:\Users\SriKriations\Favorites\Downloads\DAILY ACCOUNTS.xlsx"
wb = openpyxl.load_workbook(file_path, data_only=True)

def clean_str(val):
    if val is None:
        return ""
    return str(val).strip().upper()

# Let's inspect rows 1-20 for all sheets that seem to contain student metadata
sheets_to_inspect = ['Sheet19', 'BUS ROUTE WISE', 'Sheet18', 'Sheet17', 'Sheet20', 'Sheet14']

for name in sheets_to_inspect:
    if name not in wb.sheetnames:
        continue
    sheet = wb[name]
    print(f"\n======================================")
    print(f"SHEET: {name}")
    print(f"======================================")
    
    # print headers and first 10 rows
    max_row = min(sheet.max_row, 25)
    for r in range(1, max_row + 1):
        row_vals = [sheet.cell(row=r, column=c).value for c in range(1, min(sheet.max_column + 1, 20))]
        # only print rows that are not entirely empty
        if any(v is not None for v in row_vals):
            print(f"Row {r:2d}: {row_vals}")
