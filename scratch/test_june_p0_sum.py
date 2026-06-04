import openpyxl

wb = openpyxl.load_workbook(r"E:\accounts\RCB june 2025 26.xlsx", data_only=True)
sheet = wb['Sheet1']

cash_sum = 0
online_sum = 0
deposit_sum = 0

for r in range(2, 33):
    rec = str(sheet.cell(row=r, column=3).value or "").strip()
    cash = sheet.cell(row=r, column=6).value or 0
    online = sheet.cell(row=r, column=7).value or 0
    
    if rec == "DEPOSIT":
        deposit_sum += cash
    else:
        cash_sum += cash
        online_sum += online

print(f"Excel Rows 2-32 Sums:")
print(f"  Cash Sum   : {cash_sum}")
print(f"  Online Sum : {online_sum}")
print(f"  Deposit Sum: {deposit_sum}")
wb.close()
