import openpyxl

wb = openpyxl.load_workbook(r"E:\accounts\RCB june 2025 26.xlsx", data_only=True)
sheet = wb['Sheet1']

print("June Specific Sheets Rows:")
for r in range(1, sheet.max_row + 1):
    vals = [sheet.cell(row=r, column=c).value for c in range(1, 10)]
    print(f"Row {r:2d}: {vals}")
wb.close()
