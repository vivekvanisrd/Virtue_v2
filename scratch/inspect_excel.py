import os
import openpyxl

file_path = r"C:\Users\SriKriations\Favorites\Downloads\DAILY ACCOUNTS.xlsx"

def inspect():
    if not os.path.exists(file_path):
        print(f"File not found at: {file_path}")
        return

    print(f"File found! Size: {os.path.getsize(file_path)} bytes")
    wb = openpyxl.load_workbook(file_path, read_only=True)
    print("Sheets in workbook:", wb.sheetnames)
    
    for sheet_name in wb.sheetnames[:5]:
        print(f"\n--- Sheet: {sheet_name} ---")
        sheet = wb[sheet_name]
        # print first 10 rows
        row_count = 0
        for row in sheet.iter_rows(values_only=True):
            if row_count >= 15:
                break
            # print non-empty rows
            if any(cell is not None for cell in row):
                print(row[:12])
                row_count += 1

if __name__ == "__main__":
    inspect()
