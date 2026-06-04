import openpyxl
import json
import os
import sys
import datetime

sys.stdout.reconfigure(encoding='utf-8')

excel_dir = r"E:\accounts"
output_dir = r"J:\virtue_fb\virtue-v2\scratch\transcribed_json"
os.makedirs(output_dir, exist_ok=True)

def clean_name(name):
    if name is None:
        return ""
    name_str = str(name).strip().upper()
    while "  " in name_str:
        name_str = name_str.replace("  ", " ")
    return name_str

def clean_class(cls):
    if cls is None:
        return ""
    return str(cls).strip().upper()

def parse_date(date_val):
    if isinstance(date_val, datetime.datetime):
        return date_val.strftime('%Y-%m-%d')
    if date_val is None:
        return ""
    d_str = str(date_val).strip()
    for fmt in ('%d/%m/%Y', '%Y-%m-%d', '%d-%m-%Y', '%d/%m/%y', '%d/%m/%Y %H:%M:%S'):
        try:
            return datetime.datetime.strptime(d_str, fmt).strftime('%Y-%m-%d')
        except ValueError:
            pass
    return d_str

def make_payments(cash_val, online_val, remarks):
    payments = []
    cash = 0
    online = 0
    try:
        if cash_val:
            cash = int(float(str(cash_val).strip()))
    except ValueError:
        pass
    try:
        if online_val:
            online = int(float(str(online_val).strip()))
    except ValueError:
        pass
        
    remarks_lower = str(remarks).lower() if remarks else ""
    is_tf = "tf" in remarks_lower or "t/f" in remarks_lower or "term" in remarks_lower
    is_transport = "transport" in remarks_lower or "tp" in remarks_lower
    
    if cash > 0:
        if is_transport:
            payments.append(["CASH (TRANSPORT)", cash])
        elif is_tf:
            payments.append(["CASH (T.F)", cash])
        else:
            payments.append(["CASH", cash])
            
    if online > 0:
        if is_transport:
            payments.append(["ONLINE (TRANSPORT)", online])
        elif is_tf:
            payments.append(["ONLINE (T.F)", online])
        else:
            payments.append(["ONLINE", online])
            
    return payments

def get_row_data(sheet, row_idx, pdf_name):
    date_str = parse_date(sheet.cell(row=row_idx, column=2).value)
    rec = str(sheet.cell(row=row_idx, column=3).value or "").strip()
    if rec.endswith(".0"):
        rec = rec[:-2]
    name = clean_name(sheet.cell(row=row_idx, column=4).value)
    cls = clean_class(sheet.cell(row=row_idx, column=5).value)
    cash = sheet.cell(row=row_idx, column=6).value
    online = sheet.cell(row=row_idx, column=7).value
    remarks = sheet.cell(row=row_idx, column=9).value
    
    # Special handle for DEPOSIT
    if rec == "DEPOSIT" or name == "TO DEPOSIT":
        return {
            "date": date_str,
            "name": "TO DEPOSIT",
            "class": "UNCLEAR",
            "receipt_no": "DEPOSIT",
            "payments": [["DEPOSIT", cash or online or 0]],
            "sheet": pdf_name
        }
        
    # Handle side T/F entries
    if rec == "T/F" or "T/F" in name:
        amt = cash or online or 0
        pmt_type = "CASH (T.F)" if cash else "ONLINE (T.F)"
        return {
            "date": date_str,
            "name": name,
            "class": cls,
            "receipt_no": "T/F",
            "payments": [[pmt_type, amt]],
            "sheet": pdf_name
        }

    payments = make_payments(cash, online, remarks)
    
    return {
        "date": date_str,
        "name": name,
        "class": cls,
        "receipt_no": rec,
        "payments": payments,
        "sheet": pdf_name
    }

def compile_month(filename, sheet_name, pdf_name, page_ranges):
    file_path = os.path.join(excel_dir, filename)
    wb = openpyxl.load_workbook(file_path, data_only=True)
    sheet = wb[sheet_name]
    
    print(f"\nProcessing {filename}...")
    for page_idx, row_range in enumerate(page_ranges):
        page_data = []
        for r in row_range:
            # Skip header or empty rows (where row values are all empty)
            row_vals = [sheet.cell(row=r, column=c).value for c in range(1, sheet.max_column+1)]
            if all(x is None or str(x).strip() == "" for x in row_vals):
                continue
            # Get transaction data
            tx = get_row_data(sheet, r, pdf_name)
            page_data.append(tx)
            
        # Set row numbers sequentially on each page
        for idx, tx in enumerate(page_data):
            tx["row"] = idx + 1
            
        out_filename = f"{pdf_name[:-4].replace(' ', '_')}_p{page_idx}_i0.jpg.json"
        filepath = os.path.join(output_dir, out_filename)
        with open(filepath, 'w', encoding='utf-8') as f:
            json.dump(page_data, f, indent=2, ensure_ascii=False)
        print(f"  Generated {out_filename} with {len(page_data)} transactions.")
        
    wb.close()

# ----------------- MONTH DEFINITIONS -----------------

# June (4 pages):
# Page 0: Rows 2-26
# Page 1: Rows 27-51
# Page 2: Rows 53-82
# Page 3: Rows 83-92
compile_month("RCB june 2025 26.xlsx", "Sheet1", "RCB june 2025 26.pdf", [
    range(2, 27),
    range(27, 52),
    range(53, 83),
    range(83, 93)
])

# July (14 pages):
# Page 0: Rows 2-21
# Page 1: Rows 22-40
# Page 2: Rows 42-62
# Page 3: Rows 64-88
# Page 4: Rows 90-110
# Page 5: Rows 113-131
# Page 6: Rows 133-142
# Page 7: Rows 144-153
# Page 8: Rows 154-168
# Page 9: Rows 170-173
# Page 10: Rows 175-195
# Page 11: Rows 196-201
# Page 12: Rows 202-210
# Page 13: Rows 211-224
compile_month("RCB july 2025 26.xlsx", "Sheet1", "RCB july 2025 26.pdf", [
    range(2, 22), range(22, 41), range(42, 63), range(64, 89),
    range(90, 111), range(113, 132), range(133, 143), range(144, 154),
    range(154, 169), range(170, 174), range(175, 196), range(196, 202),
    range(202, 211), range(211, 225)
])

# August (9 pages):
# Roughly split 82 rows into 9 pages of ~9 rows each
compile_month("August_2025_Fee_Extraction.xlsx", "August_2025", "RCB August 2025 26.pdf", [
    range(2, 11), range(11, 20), range(20, 29), range(29, 38),
    range(38, 47), range(47, 56), range(56, 65), range(65, 74),
    range(74, 84)
])

# September (9 pages):
# Page 0: Rows 2-21
# Page 1: Rows 22-41
# Page 2: Rows 42-62
# Page 3: Rows 65-84
# Page 4: Rows 85-101
# Page 5: Rows 103-115
# Page 6-8: Empty summary/deposit pages
compile_month("RCB Sep 2025 26.xlsx", "Sheet1", "RCB Sep 2025 26.pdf", [
    range(2, 22), range(22, 42), range(42, 63), range(65, 85),
    range(85, 102), range(103, 116), range(0, 0), range(0, 0), range(0, 0)
])

# October (10 pages):
# Page 0: Rows 2-21
# Page 1: Rows 22-41
# Page 2: Rows 42-62
# Page 3: Rows 64-82
# Page 4: Rows 83-101
# Page 5: Rows 102-120
# Page 6: Rows 122-137
# Page 7: Rows 139-159
# Page 8: Rows 160-181
# Page 9: Rows 182-219
compile_month("RCB Oct 2025 26.2pdf.xlsx", "Sheet1", "RCB Oct 2025 26.pdf", [
    range(2, 22), range(22, 42), range(42, 63), range(64, 83),
    range(83, 102), range(102, 121), range(122, 138), range(139, 160),
    range(160, 182), range(182, 220)
])

# November (4 pages):
# Page 0: Rows 2-15
# Page 1: Rows 16-29
# Page 2: Rows 30-43
# Page 3: Rows 44-54
compile_month("November_2025_Complete_Extraction.xlsx", "November_2025", "RCB Nov 2025 26.pdf", [
    range(2, 16), range(16, 30), range(30, 44), range(44, 55)
])

# February to April (10 pages) from the parts:
def compile_feb_apr():
    print("\nProcessing Feb to April Parts...")
    parts = [
        ("Feb_Mar_Apr_2026_Extraction_Part1.xlsx", "Feb_Mar_Apr_2026"),
        ("Feb_Mar_Apr_2026_Extraction_Part2.xlsx", "Part_2"),
        ("Feb_Mar_Apr_2026_Extraction_Part3.xlsx", "Part_3"),
        ("Feb_Mar_Apr_2026_Extraction_Part4.xlsx", "Part_4"),
        ("Feb_Mar_Apr_2026_Extraction_Part5.xlsx", "Part_5")
    ]
    
    page_idx = 0
    pdf_name = "RCB Feb to April 2025 26.pdf"
    
    for filename, sheet_name in parts:
        file_path = os.path.join(excel_dir, filename)
        wb = openpyxl.load_workbook(file_path, data_only=True)
        sheet = wb[sheet_name]
        
        # Split each part into 2 pages of equal size
        total_rows = sheet.max_row
        valid_rows = []
        for r in range(2, total_rows + 1):
            row_vals = [sheet.cell(row=r, column=c).value for c in range(1, sheet.max_column+1)]
            if all(x is None or str(x).strip() == "" for x in row_vals):
                continue
            valid_rows.append(r)
            
        mid = len(valid_rows) // 2
        p1_rows = valid_rows[:mid]
        p2_rows = valid_rows[mid:]
        
        for p_rows in [p1_rows, p2_rows]:
            page_data = []
            for r in p_rows:
                tx = get_row_data(sheet, r, pdf_name)
                page_data.append(tx)
                
            for idx, tx in enumerate(page_data):
                tx["row"] = idx + 1
                
            out_filename = f"{pdf_name[:-4].replace(' ', '_')}_p{page_idx}_i0.jpg.json"
            filepath = os.path.join(output_dir, out_filename)
            with open(filepath, 'w', encoding='utf-8') as f:
                json.dump(page_data, f, indent=2, ensure_ascii=False)
            print(f"  Generated {out_filename} with {len(page_data)} transactions.")
            page_idx += 1
            
        wb.close()

compile_feb_apr()

print("\nAll months compiled successfully.")
