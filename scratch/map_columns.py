import openpyxl

file_path = r"C:\Users\SriKriations\Favorites\Downloads\DAILY ACCOUNTS.xlsx"
wb = openpyxl.load_workbook(file_path, data_only=True)

for sheet_name in wb.sheetnames:
    sheet = wb[sheet_name]
    print(f"\nSheet: {sheet_name}")
    # find all cells containing "NAME" or "AMOUNT" or "DATE" (case insensitive, stripped)
    headers = []
    max_row = min(sheet.max_row, 30) # check first 30 rows for headers
    for r in range(1, max_row + 1):
        for c in range(1, sheet.max_column + 1):
            val = sheet.cell(row=r, column=c).value
            if val and isinstance(val, str):
                val_clean = val.strip().upper()
                if "NAME" in val_clean or "DATE" in val_clean or "AMOUNT" in val_clean or "CLASS" in val_clean:
                    headers.append((r, c, val_clean))
    
    # group headers by row
    rows = {}
    for r, c, val in headers:
        if r not in rows:
            rows[r] = []
        rows[r].append((c, val))
    
    for r in sorted(rows.keys()):
        print(f"  Row {r}: {rows[r]}")
