import openpyxl

file_path = r"C:\Users\SriKriations\Favorites\Downloads\DAILY ACCOUNTS.xlsx"
wb = openpyxl.load_workbook(file_path, data_only=True)

sheets = ['Sheet19', 'Sheet18', 'Sheet17', 'Sheet20']

for s_name in sheets:
    if s_name not in wb.sheetnames:
        continue
    sheet = wb[s_name]
    print(f"\n======================================")
    print(f"SHEET: {s_name}")
    print(f"======================================")
    
    # print rows 1 to 10
    for r in range(1, 12):
        row_vals = [sheet.cell(row=r, column=c).value for c in range(1, min(sheet.max_column + 1, 15))]
        if any(v is not None for v in row_vals):
            print(f"Row {r}: {row_vals}")
