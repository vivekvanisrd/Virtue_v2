import openpyxl

wb = openpyxl.load_workbook(r"E:\accounts\sorted fee 2025 2026.xlsx", read_only=True)

for name in ["Ucban", "Feb_Mar_Apr_2026"]:
    sheet = wb[name]
    print(f"=== Sheet: {name} ===")
    
    # Read the first few rows to find the headers row
    header_row = 1
    found_headers = []
    
    for r in range(1, 10):
        row_vals = [sheet.cell(row=r, column=c).value for c in range(1, 20)]
        non_empty = [v for v in row_vals if v is not None]
        if any("Student Name" in str(v) or "Student" in str(v) for v in non_empty):
            header_row = r
            found_headers = [str(v).strip() if v is not None else "" for v in row_vals]
            break
            
    print(f"Header Row: {header_row}")
    print(f"Headers (first 15): {found_headers[:15]}")
    
    # Count rows by reading column 8 (Student Name or similar)
    name_col = 8 if name == "Ucban" else 7
    total_rows = 0
    non_empty_rows = 0
    for r in range(header_row + 1, 3000):
        val = sheet.cell(row=r, column=name_col).value
        if val is not None:
            non_empty_rows += 1
        total_rows += 1
        # Stop check if we hit 100 consecutive None
        if r > 100 and all(sheet.cell(row=x, column=name_col).value is None for x in range(r-50, r)):
            break
    print(f"Estimated non-empty rows: {non_empty_rows}\n")
