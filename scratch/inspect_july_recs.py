import openpyxl
wb = openpyxl.load_workbook(r"E:\accounts\RCB july 2025 26.xlsx", data_only=True)
sheet = wb['Sheet1']

for r in range(130, sheet.max_row + 1):
    rec = sheet.cell(row=r, column=3).value
    name = sheet.cell(row=r, column=4).value
    cash = sheet.cell(row=r, column=6).value
    online = sheet.cell(row=r, column=7).value
    remarks = sheet.cell(row=r, column=9).value
    print(f"Row {r:3d} | Rec: {rec} | Name: {name} | Cash: {cash} | Online: {online} | Remarks: {remarks}")
wb.close()
