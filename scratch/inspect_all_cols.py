import os
import openpyxl
import sys

sys.stdout.reconfigure(encoding='utf-8')
dir_path = r"E:\accounts"

files_and_sheets = [
    ("RCB june 2025 26.xlsx", "Sheet1"),
    ("RCB july 2025 26.xlsx", "Sheet1"),
    ("August_2025_Fee_Extraction.xlsx", "August_2025"),
    ("RCB Sep 2025 26.xlsx", "Sheet1"),
    ("RCB Oct 2025 26.2pdf.xlsx", "Sheet1"),
    ("November_2025_Complete_Extraction.xlsx", "November_2025"),
]

for filename, sheet_name in files_and_sheets:
    file_path = os.path.join(dir_path, filename)
    wb = openpyxl.load_workbook(file_path, data_only=True)
    sheet = wb[sheet_name]
    
    # Print all headers (Row 1)
    headers = [sheet.cell(row=1, column=c).value for c in range(1, sheet.max_column+1)]
    # Filter out None values at the end
    while headers and headers[-1] is None:
        headers.pop()
    print(f"File: {filename:35s} | Headers: {headers}")
    wb.close()
