import openpyxl
import sys

sys.stdout.reconfigure(encoding='utf-8')

wb = openpyxl.load_workbook(r"E:\accounts\sorted fee 2025 2026.xlsx", data_only=True)
sheet = wb['Feb_Mar_Apr_2026']

receipts_to_find = [5984, 6000, 5301, 5302, 5325, 5326, 5351, 5352, 5380, 5381, 5400, 5206, 5218, 5225, 5236, 5260, 5281, 5282, 5299, 5300]

print("Scanning all rows in Feb_Mar_Apr_2026 to locate the relevant receipt numbers:")
found = {}
for r in range(3, sheet.max_row + 1):
    rec_val = sheet.cell(row=r, column=6).value
    # try converting to int if it's numeric/string
    try:
        rec_int = int(float(str(rec_val).strip()))
    except ValueError:
        rec_int = str(rec_val).strip()
    
    # Check if this row is related
    # We can also check if receipt matches our list
    if isinstance(rec_int, int) and rec_int in receipts_to_find:
        found[rec_int] = (r, [sheet.cell(row=r, column=c).value for c in range(1, 13)])

for rec in sorted(list(receipts_to_find)):
    if rec in found:
        print(f"Receipt {rec}: Row {found[rec][0]} -> {found[rec][1]}")
    else:
        print(f"Receipt {rec}: NOT FOUND in sheet")

wb.close()
