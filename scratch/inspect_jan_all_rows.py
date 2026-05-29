import openpyxl
import datetime

wb = openpyxl.load_workbook(r"E:\accounts\sorted fee 2025 2026.xlsx", data_only=True)
sheet = wb['Feb_Mar_Apr_2026']

cash_total = 0
online_total = 0
matching_rows = []

for r in range(3, sheet.max_row + 1):
    month_val = sheet.cell(row=r, column=4).value
    date_val = sheet.cell(row=r, column=5).value
    cash = sheet.cell(row=r, column=9).value or 0
    online = sheet.cell(row=r, column=10).value or 0
    
    is_jan = False
    # Check if month is Jan 2026
    if isinstance(month_val, datetime.datetime) and month_val.year == 2026 and month_val.month == 1:
        is_jan = True
    elif isinstance(month_val, str) and "jan" in month_val.lower():
        is_jan = True
        
    if is_jan:
        cash_total += cash
        online_total += online
        matching_rows.append(r)

print(f"Total January rows found: {len(matching_rows)}")
print(f"Row range: {min(matching_rows) if matching_rows else None} to {max(matching_rows) if matching_rows else None}")
print(f"Total Cash: {cash_total}, Total Online: {online_total}, Grand Total: {cash_total + online_total}")

wb.close()
