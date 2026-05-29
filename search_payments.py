import openpyxl
import sys
import warnings
from datetime import datetime

# Suppress openpyxl user warnings about cell formats
warnings.filterwarnings("ignore", category=UserWarning, module="openpyxl")

FILE_PATH = r"C:\Users\SriKriations\Favorites\Downloads\DAILY ACCOUNTS.xlsx"

def clean_str(val):
    if val is None:
        return ""
    return str(val).strip()

def matches_query(name, query):
    if not name or not query:
        return False
    # normalize spaces and do case-insensitive substring match
    n = " ".join(name.upper().split())
    q = " ".join(query.upper().split())
    return q in n

def format_date(val):
    if val is None:
        return ""
    if isinstance(val, datetime):
        return val.strftime("%Y-%m-%d")
    return str(val).split()[0]

def search_excel(query):
    wb = openpyxl.load_workbook(FILE_PATH, data_only=True)
    results = []
    
    for sheet_name in wb.sheetnames:
        sheet = wb[sheet_name]
        
        # We will scan all cells to find "NAME" header columns
        max_row = min(sheet.max_row, 100) # search first 100 rows for headers
        table_headers = []
        
        for r in range(1, max_row + 1):
            row_vals = [sheet.cell(row=r, column=c).value for c in range(1, sheet.max_column + 1)]
            
            for c_idx, val in enumerate(row_vals):
                c = c_idx + 1
                val_str = clean_str(val).upper()
                
                # Identify Name columns
                if ("NAME" in val_str or "STUDENT" in val_str) and "ABOUT" not in val_str and "EXPENDITURE" not in val_str:
                    # Found a name column! Now search for closest date, class, and amount columns
                    name_col = c
                    date_col = None
                    class_col = None
                    amount_cols = []
                    
                    # We look for Date, Class, and Amount in the same row
                    for other_c_idx, other_val in enumerate(row_vals):
                        other_c = other_c_idx + 1
                        other_val_str = clean_str(other_val).upper()
                        
                        if not other_val_str:
                            continue
                        
                        # Date should be to the left of the Name column (or very close)
                        if "DATE" in other_val_str or other_val_str == "DT":
                            if date_col is None or abs(other_c - name_col) < abs(date_col - name_col):
                                date_col = other_c
                                
                        # Class should be near Name (usually to the right)
                        if "CLASS" in other_val_str or other_val_str == "CL":
                            if class_col is None or abs(other_c - name_col) < abs(class_col - name_col):
                                class_col = other_c
                                
                        # Amounts should be to the right of Name, within 6 columns
                        if any(x in other_val_str for x in ["AMOUNT", "CASH", "ONLINE", "TOTAL", "FEE", "COLLECTED", "PAYMENT"]):
                            if other_c > name_col and (other_c - name_col) <= 6:
                                amount_cols.append((other_c, other_val_str))
                    
                    table_headers.append({
                        "header_row": r,
                        "name_col": name_col,
                        "date_col": date_col,
                        "class_col": class_col,
                        "amount_cols": amount_cols
                    })
        
        # Now parse each table block we found on this sheet
        for tb in table_headers:
            name_col = tb["name_col"]
            date_col = tb["date_col"]
            class_col = tb["class_col"]
            amount_cols = tb["amount_cols"]
            start_row = tb["header_row"] + 1
            
            # Scan down from start_row
            for r in range(start_row, sheet.max_row + 1):
                name_val = clean_str(sheet.cell(row=r, column=name_col).value)
                
                # Stop parsing if we hit a blank name or a total/summary row
                name_upper = name_val.upper()
                if not name_val or any(x in name_upper for x in ["TOTAL", "GRAND TOTAL", "SUBTOTAL", "EXPENDITURE", "MONTHLY", "WEEK"]):
                    break
                
                # Check for query match
                if matches_query(name_val, query):
                    # Get date
                    date_val = format_date(sheet.cell(row=r, column=date_col).value) if date_col else ""
                    
                    # Get class
                    class_val = clean_str(sheet.cell(row=r, column=class_col).value) if class_col else ""
                    
                    # Get amounts
                    payments = []
                    for ac, label in amount_cols:
                        amt_val = sheet.cell(row=r, column=ac).value
                        if amt_val is not None:
                            try:
                                # Convert to float/int if possible
                                amt_num = float(amt_val)
                                if amt_num != 0:
                                    payments.append((label, amt_num))
                            except ValueError:
                                if clean_str(amt_val):
                                    payments.append((label, clean_str(amt_val)))
                    
                    results.append({
                        "sheet": sheet_name,
                        "date": date_val,
                        "name": name_val,
                        "class": class_val,
                        "payments": payments,
                        "row": r
                    })
                    
    return results

def main():
    if len(sys.argv) < 2:
        print("Usage: python search_payments.py <student_name>")
        sys.exit(1)
        
    query = " ".join(sys.argv[1:])
    print(f"\nSearching for payments matching: '{query}'...")
    print("-" * 80)
    
    results = search_excel(query)
    
    if not results:
        print("No payment records found matching that name.")
        return
        
    print(f"Found {len(results)} records:")
    print(f"{'Sheet':<12} | {'Date':<10} | {'Student Name':<28} | {'Class':<8} | {'Payments':<30}")
    print("-" * 96)
    
    total_amount = 0.0
    for r in results:
        payments_str = ", ".join([f"{val} ({label})" if isinstance(val, (int, float)) else f"{val} ({label})" for label, val in r["payments"]])
        # Add numeric amounts to total
        for label, val in r["payments"]:
            if isinstance(val, (int, float)):
                total_amount += val
        
        print(f"{r['sheet']:<12} | {r['date']:<10} | {r['name']:<28} | {r['class']:<8} | {payments_str:<30}")
        
    print("-" * 96)
    print(f"Total Numeric Amount Extracted: {total_amount:,.2f}")

if __name__ == "__main__":
    main()
