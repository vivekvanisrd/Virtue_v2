import openpyxl
import sys
import re
import json
import warnings
import os
from datetime import datetime

# Suppress openpyxl formatting warnings
warnings.filterwarnings("ignore", category=UserWarning, module="openpyxl")

SORTED_FEE_PATH = r"E:\accounts\sorted fee 2025 2026.xlsx"
OUTPUT_JSON_PATH = r"J:\virtue_fb\virtue-v2\public\student-directory\student_data.json"

def clean_str(val):
    if val is None:
        return ""
    return str(val).strip()

def clean_name(name):
    if not name:
        return ""
    s = str(name).strip().upper()
    s = re.sub(r'\s+', ' ', s)
    s = re.sub(r'\.+', '.', s)
    s = re.sub(r'\.(\w)', r'. \1', s)
    s = re.sub(r'\s+\.', '.', s)
    s = re.sub(r'\s+', ' ', s).strip()
    return s

def normalize_class(c_str):
    c = clean_str(c_str).upper()
    c = "".join(c.split()) # remove spaces
    if not c:
        return ""
    
    if c in ["NUR", "NURSERY", "N", "PLAY", "PLAYGROUP", "PG"]:
        return "NUR"
    if c in ["LKG", "PP-1", "PP1", "PPI", "L.KG", "PP-IA", "PPIA"]:
        return "LKG"
    if c in ["UKG", "PP-2", "PP2", "PPII", "U.KG", "PP-IIA", "PPIIA"]:
        return "UKG"
    
    for num, term in [("10", "10TH"), ("9", "9TH"), ("8", "8TH"), ("7", "7TH"), ("6", "6TH"), ("5", "5TH"), ("4", "4TH"), ("3", "3RD"), ("2", "2ND"), ("1", "1ST")]:
        if num in c:
            return term
            
    return c

def format_date(val):
    if val is None:
        return ""
    if isinstance(val, datetime):
        return val.strftime("%Y-%m-%d")
    val_str = str(val).strip()
    for fmt in ("%d/%m/%Y", "%Y-%m-%d", "%d-%m-%Y"):
        try:
            return datetime.strptime(val_str, fmt).strftime("%Y-%m-%d")
        except ValueError:
            pass
    return val_str.split()[0]

def main():
    print(f"Loading payments from final sorted Excel sheet: {SORTED_FEE_PATH} (Ucban sheet only)...")
    wb_sorted = openpyxl.load_workbook(SORTED_FEE_PATH, data_only=True)
    if "Ucban" not in wb_sorted.sheetnames:
        print("Error: Ucban sheet not found in workbook.")
        sys.exit(1)
        
    sheet_sorted = wb_sorted["Ucban"]
    
    grouped_payments = {}
    total_rows_read = 0
    total_total_column_sum = 0.0
    
    for r in range(4, sheet_sorted.max_row + 1):
        name_val = sheet_sorted.cell(row=r, column=8).value
        if not name_val:
            continue
            
        name_str = str(name_val).strip()
        name_upper = name_str.upper()
        if any(x in name_upper for x in ["DEPOSIT", "EXPENDITURE", "TOTAL", "GRAND TOTAL", "CONTRA", "CEMENT", "DIESEL"]):
            continue
            
        norm_name = clean_name(name_str)
        class_val = clean_str(sheet_sorted.cell(row=r, column=9).value)
        date_val = sheet_sorted.cell(row=r, column=6).value
        date_str = format_date(date_val)
        
        receipt_val = sheet_sorted.cell(row=r, column=7).value
        receipt_str = clean_str(receipt_val)
        if receipt_str.endswith(".0"):
            receipt_str = receipt_str[:-2]
        receipt_str = re.sub(r'\s+', '', receipt_str)
        
        branch_val = clean_str(sheet_sorted.cell(row=r, column=4).value)
        branch = branch_val if branch_val and branch_val != "None" else "RCB"
        
        total_val = sheet_sorted.cell(row=r, column=12).value
        try:
            total = float(total_val) if total_val is not None else 0.0
        except ValueError:
            total = 0.0
            
        if total == 0:
            continue
            
        cash_val = sheet_sorted.cell(row=r, column=10).value
        try:
            cash = float(cash_val) if cash_val is not None else 0.0
        except ValueError:
            cash = 0.0
            
        online_val = sheet_sorted.cell(row=r, column=11).value
        try:
            online = float(online_val) if online_val is not None else 0.0
        except ValueError:
            online = 0.0
            
        remarks_val = clean_str(sheet_sorted.cell(row=r, column=13).value)
        
        pay_items = []
        if cash == 0 and online == 0:
            label = f"PAYMENT ({remarks_val})" if remarks_val else "PAYMENT"
            pay_items.append((label, total))
        elif cash > 0 and online == 0:
            label = f"CASH ({remarks_val})" if remarks_val else "CASH"
            pay_items.append((label, total))
        elif online > 0 and cash == 0:
            label = f"ONLINE ({remarks_val})" if remarks_val else "ONLINE"
            pay_items.append((label, total))
        else: # both cash and online are non-zero
            c_label = f"CASH ({remarks_val})" if remarks_val else "CASH"
            o_label = f"ONLINE ({remarks_val})" if remarks_val else "ONLINE"
            pay_items.append((c_label, cash))
            pay_items.append((o_label, total - cash))
            
        payment_record = {
            "sheet": "sorted fee 2025 2026.xlsx",
            "date": date_str,
            "name": name_str,
            "class": class_val,
            "receipt_no": receipt_str,
            "payments": pay_items,
            "row": r
        }
        
        total_rows_read += 1
        total_total_column_sum += sum(item[1] for item in pay_items)
        
        if norm_name not in grouped_payments:
            grouped_payments[norm_name] = {
                "name_in_sheet": name_str,
                "class_in_sheet": class_val,
                "branch": branch,
                "payments": []
            }
        grouped_payments[norm_name]["payments"].append(payment_record)
        
    print(f"Read {total_rows_read} rows from Ucban. Total payments sum parsed: Rs. {total_total_column_sum:,.2f}")
    print(f"Grouped into {len(grouped_payments)} unique students.")
    
    # Construct final student database
    output_data = []
    total_payments_sum = 0.0
    total_assigned_payments = 0
    
    for student_key, entry in grouped_payments.items():
        profile_class = entry["class_in_sheet"]
        branch = entry["branch"]
        
        student_payments = entry["payments"]
        student_total_paid = 0.0
        for p in student_payments:
            for label, val in p["payments"]:
                if isinstance(val, (int, float)):
                    student_total_paid += val
                    
        total_assigned_payments += len(student_payments)
        total_payments_sum += student_total_paid
        
        output_data.append({
            "id": "",
            "name": clean_name(entry["name_in_sheet"]),
            "class": profile_class,
            "normalized_class": normalize_class(profile_class),
            "section": "",
            "branch": branch,
            "father": "",
            "mother": "",
            "f_phone": "",
            "m_phone": "",
            "aadhar": "",
            "address": "",
            "bus_route": "",
            "pickup_point": "",
            "committed_fee": 0.0,
            "source": "sorted fee 2025 2026.xlsx (Ucban sheet)",
            "total_paid": student_total_paid,
            "payments": student_payments
        })
        
    print(f"Total payments matched and assigned: {total_assigned_payments}")
    print(f"Grand total sum of collections on dashboard: Rs. {total_payments_sum:,.2f}")
    
    # Save student_data.json
    print(f"Writing database to {OUTPUT_JSON_PATH}...")
    with open(OUTPUT_JSON_PATH, "w", encoding="utf-8") as f:
        json.dump(output_data, f, indent=2, ensure_ascii=False)
        
    # Save student_data.js
    OUTPUT_JS_PATH = OUTPUT_JSON_PATH.replace(".json", ".js")
    print(f"Writing database to {OUTPUT_JS_PATH}...")
    with open(OUTPUT_JS_PATH, "w", encoding="utf-8") as f:
        f.write("window.studentsData = ")
        json.dump(output_data, f, indent=2, ensure_ascii=False)
        f.write(";\n")
        
    print("Database built successfully!")

if __name__ == "__main__":
    main()
